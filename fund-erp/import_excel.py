# -*- coding: utf-8 -*-
"""푸른노무법인 통합 엑셀 → funds / sites / site_histories / agencies 초기 이관
사용: python import_excel.py ["엑셀경로"]
- 2025지역공근사업장 / 2026지역공근사업장 시트를 읽어 연도 스냅샷으로 적재
- 연번 규칙: datetime(월=호수) 또는 "N-M"=충남 N호, "경N-M"=경기 N호
- 개인정보 원칙(0단계): 주민번호 등 민감정보는 애초에 읽지 않음. 담당자 업무연락처만 저장.
"""
import sys, os, re, json
from datetime import datetime
import openpyxl
from db import connect, init_db, audit

DEFAULT_XLSX = r"C:\Users\fair0\Downloads\푸른노무법인 (4).xlsx"

# 충남 호수별 출연 지자체 (폴더 실사례 기준)
CN_CITIES = {
    1: ["예산군", "공주시", "보령시"], 2: ["아산시", "보령시"], 3: ["공주시", "서천군", "태안군"],
    4: ["청양군"], 5: ["부여군", "홍성군"], 6: ["서산시"], 7: ["서천군", "부여군", "논산시"],
    8: ["당진시"], 9: ["천안시"], 10: ["공주시", "예산군", "금산군"], 11: ["태안군", "서산시"],
    12: ["보령시", "홍성군"],
}


def fund_no(v):
    """연번 → (지역, 호수) 또는 None"""
    if isinstance(v, datetime):
        return ("충남", v.month)
    s = str(v or "").strip()
    m = re.match(r"^경\s*(\d+)\s*[-–]", s)
    if m:
        return ("경기", int(m.group(1)))
    m = re.match(r"^(\d+)\s*[-–]", s)
    if m:
        return ("충남", int(m.group(1)))
    return None


def seq_label(v, region, no):
    if isinstance(v, datetime):
        return f"{no}-{v.day}"
    return str(v).strip()


def norm_bizno(v):
    d = re.sub(r"[^0-9]", "", str(v or ""))[:10]
    if len(d) == 10:
        return f"{d[:3]}-{d[3:5]}-{d[5:]}"
    return str(v or "").strip()


def load_sheet(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(rows[:5]) if r and r[0] and "연번" in str(r[0]))
    return [r for r in rows[hdr_i + 1:] if r and not all(x is None or str(x).strip() == "" for x in r)]


def ensure_agency(con, name, atype):
    row = con.execute("SELECT agency_id FROM agencies WHERE name=?", (name,)).fetchone()
    if row:
        return row[0]
    n = con.execute("SELECT COUNT(*) FROM agencies").fetchone()[0]
    aid = f"AG-{n + 1:04d}"
    con.execute("INSERT INTO agencies(agency_id,name,agency_type) VALUES(?,?,?)", (aid, name, atype))
    return aid


def link_agency(con, fund_id, agency_id, role):
    con.execute(
        "INSERT OR IGNORE INTO fund_agencies(fund_id,agency_id,role) VALUES(?,?,?)",
        (fund_id, agency_id, role),
    )


def ensure_fund(con, region, no):
    short = f"{region} {no}호"
    row = con.execute("SELECT fund_id FROM funds WHERE short_name=?", (short,)).fetchone()
    if row:
        return row[0]
    n = con.execute("SELECT COUNT(*) FROM funds").fetchone()[0]
    fid = f"FUND-{n + 1:04d}"
    name = f"{'충남공동근로복지기금' if region == '충남' else '경기공동근로복지기금'}{no}호"
    if region == "충남" and no <= 3:
        name = f"더행복한충남공동근로복지기금{'' if no == 1 else str(no) + '호'}"
    con.execute(
        "INSERT INTO funds(fund_id,name,short_name,fund_type,region) VALUES(?,?,?,?,?)",
        (fid, name, short, "공동", region),
    )
    audit(con, "funds", fid, "create", after=name)
    # 관계기관 자동 연결
    pnp = ensure_agency(con, "푸른노무법인", "수탁법인")
    kcomwel = ensure_agency(con, "근로복지공단", "공단")
    link_agency(con, fid, pnp, "수탁기관")
    link_agency(con, fid, kcomwel, "지원기관")
    if region == "충남":
        do = ensure_agency(con, "충청남도", "지자체")
        link_agency(con, fid, do, "위탁기관")
        link_agency(con, fid, do, "출연지자체")
        for city in CN_CITIES.get(no, []):
            cid = ensure_agency(con, city, "지자체")
            link_agency(con, fid, cid, "출연지자체")
    else:
        do = ensure_agency(con, "경기도", "지자체")
        link_agency(con, fid, do, "위탁기관")
        link_agency(con, fid, do, "출연지자체")
    return fid


def import_year(con, wb, sheet, year):
    data = load_sheet(wb, sheet)
    added_sites = updated = snaps = skipped = 0
    for r in data:
        fn = fund_no(r[0])
        if not fn:
            skipped += 1
            continue
        region, no = fn
        fid = ensure_fund(con, region, no)
        name = str(r[1] or "").strip()
        biz = norm_bizno(r[7] if len(r) > 7 else "")
        if not name:
            skipped += 1
            continue
        # 사업자번호 우선 매칭, 없으면 (기금, 상호) 매칭
        row = None
        if biz:
            row = con.execute("SELECT site_id FROM sites WHERE biz_no=?", (biz,)).fetchone()
        if not row:
            row = con.execute(
                "SELECT site_id FROM sites WHERE fund_id=? AND name=?", (fid, name)
            ).fetchone()
        contacts = []
        cname = str(r[8] or "").strip() if len(r) > 8 else ""
        if cname:
            contacts.append({
                "name": cname,
                "position": str(r[9] or "").strip() if len(r) > 9 else "",
                "phone": str(r[10] or "").strip() if len(r) > 10 else "",
                "mobile": str(r[11] or "").strip() if len(r) > 11 else "",
                "email": str(r[12] or "").strip() if len(r) > 12 else "",
                "isPrimary": True,
            })
        if row:
            sid = row[0]
            con.execute(
                "UPDATE sites SET name=?, ceo=?, address=?, biz_type=?, contacts=?, "
                "updated_at=datetime('now','localtime') WHERE site_id=?",
                (name, str(r[2] or "").strip(), str(r[3] or "").strip(),
                 str(r[5] or "").strip(), json.dumps(contacts, ensure_ascii=False), sid),
            )
            updated += 1
        else:
            n = con.execute("SELECT COUNT(*) FROM sites").fetchone()[0]
            sid = f"SITE-{n + 1:04d}"
            con.execute(
                "INSERT INTO sites(site_id,fund_id,seq_label,name,biz_no,ceo,address,biz_type,contacts)"
                " VALUES(?,?,?,?,?,?,?,?,?)",
                (sid, fid, seq_label(r[0], region, no), name, biz,
                 str(r[2] or "").strip(), str(r[3] or "").strip(), str(r[5] or "").strip(),
                 json.dumps(contacts, ensure_ascii=False)),
            )
            audit(con, "sites", sid, "create", after=name)
            added_sites += 1
        emp = r[4] if len(r) > 4 else None
        try:
            emp = int(emp) if emp is not None and str(emp).strip() != "" else None
        except (TypeError, ValueError):
            emp = None
        wrep = str(r[6] or "").strip() if len(r) > 6 else ""
        con.execute(
            "INSERT INTO site_histories(site_id,year,employees,worker_rep)"
            " VALUES(?,?,?,?) ON CONFLICT(site_id,year) DO UPDATE SET"
            " employees=excluded.employees, worker_rep=excluded.worker_rep",
            (sid, year, emp, wrep),
        )
        snaps += 1
    return added_sites, updated, snaps, skipped


def seed_forms(con):
    forms = [
        ("F-07-20250414", "별지 제7호서식", "설립인가신청서", "2025-04-14"),
        ("F-10-20210609", "별지 제10호서식", "기본재산 총액 변경 내용 보고서", "2021-06-09"),
        ("F-11-20210609", "별지 제11호서식", "정관변경 인가신청서", "2021-06-09"),
        ("F-14-20210609", "별지 제14호서식", "해산통지서", "2021-06-09"),
        ("F-15-20251001", "별지 제15호서식", "운영상황보고서(4쪽)", "2025-10-01"),
    ]
    for fv_id, no, name, rev in forms:
        con.execute(
            "INSERT OR IGNORE INTO form_versions(form_version_id,form_no,name,revised_date,"
            "effective_from,source) VALUES(?,?,?,?,?,?)",
            (fv_id, no, name, rev, rev, "국가법령정보센터(고용노동부령 제453호)"),
        )


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        print(f"엑셀 파일 없음: {xlsx}")
        sys.exit(1)
    init_db()
    con = connect()
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    total = {}
    for sheet, year in [("2025지역공근사업장", 2025), ("2026지역공근사업장", 2026)]:
        if sheet in wb.sheetnames:
            total[year] = import_year(con, wb, sheet, year)
    seed_forms(con)
    audit(con, "system", "import", "import", after=os.path.basename(xlsx))
    con.commit()
    for year, (a, u, s, sk) in total.items():
        print(f"[{year}] 사업장 신규 {a} / 갱신 {u} / 스냅샷 {s} / 건너뜀 {sk}")
    n_funds = con.execute("SELECT COUNT(*) FROM funds").fetchone()[0]
    n_sites = con.execute("SELECT COUNT(*) FROM sites").fetchone()[0]
    n_ag = con.execute("SELECT COUNT(*) FROM agencies").fetchone()[0]
    print(f"기금 {n_funds} / 사업장 {n_sites} / 기관 {n_ag}")
    con.close()


if __name__ == "__main__":
    main()
