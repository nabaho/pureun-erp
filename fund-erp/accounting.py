# -*- coding: utf-8 -*-
"""회계·결산 (M11·M12·M13) — 통장 파싱·자동분개 제안(승인형)·시산표·재무제표·별지15호 매핑
기존 기금결산자동화(bank_parser·journal_engine·config)의 로직을 self-contained로 이식.
자동은 '제안'만 — 확정은 담당자 승인. 이자수익 준비금 전입·감사의견·사용가능액 자동확정 금지(§8).
"""
import re, hashlib
from datetime import datetime
from openpyxl import load_workbook

# ── 계정과목 (config.py 계승) ──
ACCOUNT_CHART = {
    "현금성자산": "자산", "정기예금": "자산", "근로자대부금": "자산",
    "고유목적사업준비금1": "부채", "고유목적사업준비금2": "부채",
    "기본재산": "자본", "이월잉여금": "자본",
    "이자수익": "수익", "고유목적사업준비금1전입수입": "수익", "고유목적사업준비금2전입수입": "수익",
    "경조사비": "비용", "동호회비": "비용", "체육문화비": "비용", "장학금": "비용",
    "생활안정자금": "비용", "기타복지비": "비용",
    "지급수수료": "비용", "사무용품비": "비용", "기타관리비": "비용",
    "고유목적사업준비금전입액": "비용",
}
# ── 자동분류 규칙 (키워드 → 차/대) ──
RULES = [
    {"kw": ["출연", "출연금", "기금출연"], "dir": "입금", "d": "현금성자산", "c": "기본재산", "basic": 1},
    {"kw": ["이자", "예금이자", "보통예금이자"], "dir": "입금", "d": "현금성자산", "c": "이자수익"},
    {"kw": ["대부상환", "대부금상환", "상환", "대출상환"], "dir": "입금", "d": "현금성자산", "c": "근로자대부금"},
    {"kw": ["수수료", "지급수수료", "이체수수료", "인터넷뱅킹"], "dir": "출금", "d": "지급수수료", "c": "현금성자산"},
    {"kw": ["대부", "대부금", "대출", "대여금", "근로자대부"], "dir": "출금", "d": "근로자대부금", "c": "현금성자산"},
    {"kw": ["경조", "축의", "조의", "결혼", "장례", "부의", "연말선물"], "dir": "출금", "d": "경조사비", "c": "현금성자산"},
    {"kw": ["동호회", "동아리", "상품권", "기프트카드", "영화"], "dir": "출금", "d": "동호회비", "c": "현금성자산"},
    {"kw": ["체육", "문화", "레크레이션", "체육대회", "워크샵", "야유회"], "dir": "출금", "d": "체육문화비", "c": "현금성자산"},
    {"kw": ["장학", "학자금", "교육비"], "dir": "출금", "d": "장학금", "c": "현금성자산"},
    {"kw": ["생활안정", "생활자금", "긴급자금"], "dir": "출금", "d": "생활안정자금", "c": "현금성자산"},
    {"kw": ["커피", "원두", "우유", "식빵", "식대", "음료", "주류", "간식", "떡", "다과", "계란"], "dir": "출금", "d": "기타복지비", "c": "현금성자산"},
    {"kw": ["사무", "복사", "인쇄", "문구"], "dir": "출금", "d": "사무용품비", "c": "현금성자산"},
    {"kw": ["정기예금", "적금", "예금가입"], "dir": "출금", "d": "정기예금", "c": "현금성자산"},
]
TRANSFER_KW = ["대체", "계좌이체", "내계좌", "자행이체", "기본재산계좌", "보통재산계좌"]
CANCEL_KW = ["취소", "정정", "반제", "환입"]


def _num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = re.sub(r"[^0-9\-]", "", str(v))
    return int(s) if s and s not in ("-",) else 0


def _date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s.split()[0] if " " in s else s, fmt.split()[0]).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s[:10]


def _load_rows(raw_bytes):
    """xlsx(openpyxl) / xls(xlrd) 자동 판별 → 2차원 값 리스트"""
    import io
    if raw_bytes[:4] == b"PK\x03\x04":                      # xlsx (zip)
        wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    if raw_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # xls (OLE2/BIFF)
        try:
            import xlrd
        except ImportError:
            raise ValueError(".xls 파일은 xlrd 설치 필요 (pip install xlrd)")
        wb = xlrd.open_workbook(file_contents=raw_bytes)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                ct = ws.cell_type(r, c)
                if ct == xlrd.XL_CELL_DATE:
                    try:
                        y, mo, dd, *_ = xlrd.xldate_as_tuple(v, wb.datemode)
                        v = "%04d-%02d-%02d" % (y, mo, dd)
                    except Exception:
                        pass
                row.append(v)
            rows.append(row)
        return rows
    raise ValueError("지원하지 않는 파일 형식 (xlsx/xls만 가능)")


def parse_bank_excel(raw_bytes):
    """엑셀 bytes(xlsx/xls) → 표준 거래 리스트. 헤더 자동 탐지, 적요=텍스트 컬럼 결합."""
    rows = _load_rows(raw_bytes)
    KW = {"date": ["거래일", "거래일시", "일자", "날짜", "거래일자"],
          "dep": ["입금", "맡기신", "입금액", "입금금액"],
          "wd": ["출금", "찾으신", "출금액", "출금금액"],
          "bal": ["잔액", "거래후잔액", "잔고"],
          # 적요로 쓸 텍스트 컬럼 (여러 개면 모두 결합)
          "desc": ["적요", "내용", "거래내용", "기재", "비고", "거래구분", "보낸분", "받는분", "메모"]}
    hi, cmap, desc_cols = None, {}, []
    for i, r in enumerate(rows[:30]):
        cells = [str(x or "").strip() for x in r]
        di = next((j for j, c in enumerate(cells) if any(k in c for k in KW["date"])), None)
        if di is None:
            continue
        cmap = {"date": di}
        for key in ("dep", "wd", "bal"):
            j = next((j for j, c in enumerate(cells) if any(k in c for k in KW[key])), None)
            if j is not None:
                cmap[key] = j
        desc_cols = [j for j, c in enumerate(cells) if any(k in c for k in KW["desc"])]
        hi = i
        break
    if hi is None:
        raise ValueError("거래내역 헤더(거래일)를 찾을 수 없습니다")
    txns = []
    for r in rows[hi + 1:]:
        if not r or cmap["date"] >= len(r):
            continue
        dv = r[cmap["date"]]
        if dv is None or str(dv).strip() == "":
            continue
        d = _date(dv)
        if not re.match(r"\d{4}-\d{2}-\d{2}", d):
            continue
        dep = _num(r[cmap["dep"]]) if "dep" in cmap and cmap["dep"] < len(r) else 0
        wd = _num(r[cmap["wd"]]) if "wd" in cmap and cmap["wd"] < len(r) else 0
        if dep == 0 and wd == 0:
            continue
        desc = " ".join(str(r[j]).strip() for j in desc_cols
                        if j < len(r) and r[j] not in (None, "")).strip()
        bal = _num(r[cmap["bal"]]) if "bal" in cmap and cmap["bal"] < len(r) else 0
        txns.append({"date": d, "desc": desc, "deposit": dep, "withdrawal": wd, "balance": bal})
    txns.sort(key=lambda t: t["date"])
    return txns


def file_hash(raw_bytes):
    return hashlib.sha256(raw_bytes).hexdigest()


def flag_of(desc):
    d = desc.replace(" ", "")
    if any(k in d for k in TRANSFER_KW):
        return "transfer"
    if any(k in d for k in CANCEL_KW):
        return "cancel"
    return ""


# 적요에서 거래처·핵심어 추출 시 걸러낼 은행 상투어
_NOISE = {"출금", "입금", "이체", "인터넷", "전자금융", "FBS", "현금", "결산", "세금",
          "대체", "자동", "은행", "통장", "근로복지기금", "복지기금", "공동", "사내", "법인",
          "주식회사", "이자세금", "지급", "수취", "송금"}


def extract_keyword(desc):
    """적요 → 학습 규칙 키워드 (가장 긴 비상투어 토큰, 보통 거래처명)"""
    toks = re.findall(r"[가-힣A-Za-z]{2,}", desc or "")
    cand = [t for t in toks if t not in _NOISE]
    if not cand:
        cand = toks
    return max(cand, key=len) if cand else ""


def propose_journal(txn, learned=None):
    """단일 거래 → 분개 제안 (확정 아님). learned 규칙을 내장 규칙보다 우선 적용."""
    desc_raw = txn["desc"]
    desc = desc_raw.lower()
    direction = "입금" if txn["deposit"] > 0 else "출금"
    amount = txn["deposit"] or txn["withdrawal"]
    flag = flag_of(desc_raw)
    if flag == "transfer":
        return {"debit": "계좌대체", "credit": "계좌대체", "amount": amount,
                "rule": "계좌간이체(수입·지출 제외)", "auto": True, "basic": 0, "flag": "transfer"}
    # 학습 규칙 우선 (같은 거래처·적요를 담당자가 전에 분류한 경우)
    for lr in (learned or []):
        if lr.get("direction") and lr["direction"] != direction:
            continue
        kw = lr["keyword"]
        if kw and kw in desc_raw:
            return {"debit": lr["debit"], "credit": lr["credit"], "amount": amount,
                    "rule": f"학습:{kw}", "auto": True,
                    "basic": 1 if lr["debit"] == "기본재산" or lr["credit"] == "기본재산" else 0,
                    "flag": flag}
    for rule in RULES:
        if rule["dir"] != direction:
            continue
        if any(k in desc for k in rule["kw"]):
            return {"debit": rule["d"], "credit": rule["c"], "amount": amount,
                    "rule": "·".join(rule["kw"][:2]), "auto": True, "basic": rule.get("basic", 0), "flag": flag}
    # 미분류
    return {"debit": "현금성자산" if direction == "입금" else "?", "amount": amount,
            "credit": "?" if direction == "입금" else "현금성자산",
            "rule": "미분류", "auto": False, "basic": 0, "flag": flag}


def trial_balance(lines):
    """분개 라인 [{side,account,amount}] → 계정별 차/대 합계·잔액"""
    acc = {}
    for ln in lines:
        a = acc.setdefault(ln["account"], {"debit": 0, "credit": 0})
        a["debit" if ln["side"] == "D" else "credit"] += ln["amount"]
    out = {}
    for name, v in acc.items():
        typ = ACCOUNT_CHART.get(name, "비용")
        d, c = v["debit"], v["credit"]
        if typ in ("자산", "비용"):
            out[name] = {"type": typ, "debit": d, "credit": c, "bal_d": max(d - c, 0), "bal_c": max(c - d, 0)}
        else:
            out[name] = {"type": typ, "debit": d, "credit": c, "bal_d": max(d - c, 0), "bal_c": max(c - d, 0)}
    return out


def financial_statements(tb, prev=None):
    """시산표 → 재무제표 요약 (전기이월 prev 반영)"""
    prev = prev or {}

    def cb(a):
        return tb.get(a, {}).get("bal_c", 0)

    def db(a):
        return tb.get(a, {}).get("bal_d", 0)

    interest = cb("이자수익")
    purpose_exp = sum(db(a) for a in ["경조사비", "동호회비", "체육문화비", "장학금", "생활안정자금", "기타복지비"])
    admin_exp = sum(db(a) for a in ["지급수수료", "사무용품비", "기타관리비"])
    cash = db("현금성자산") + prev.get("cash", 0)
    savings = db("정기예금") + prev.get("savings", 0)
    loan = db("근로자대부금") + prev.get("loan", 0)
    basic = cb("기본재산") + prev.get("basic", 0)
    total_assets = cash + savings + loan
    net_income = interest - purpose_exp - admin_exp
    retained = prev.get("retained", 0) + net_income
    return {
        "interest": interest, "purpose_exp": purpose_exp, "admin_exp": admin_exp,
        "cash": cash, "savings": savings, "loan": loan, "basic": basic,
        "total_assets": total_assets, "net_income": net_income, "retained": retained,
        "total_equity": basic + retained,
        # 별지 제15호 매핑 (천원 단위)
        "form15": {
            "기본재산총액": basic, "당기운용수익금": interest,
            "근로자대부": loan, "목적사업비": purpose_exp, "운영비": admin_exp,
        },
    }
