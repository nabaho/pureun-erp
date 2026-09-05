# -*- coding: utf-8 -*-
"""과거자료 폴더 스캔 → 기금×서류종류 보유 매트릭스 (기존 기금 데이터화 2차)
- 읽기 전용: 파일을 옮기지 않고 경로만 수집
- --register: 분류된 파일을 documents 테이블에 문서철로 등록 (기금별·종류별)
사용: python scan_archive.py [--register]
출력: scan/스캔리포트.xlsx, scan/scan_report.json
"""
import os, sys, re, json
from collections import defaultdict
from datetime import datetime
from db import connect, init_db, audit

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 폴더 재편(2026-07-16): 루트 = 00_표준서식 / 01_계획·설계 / 02_프로그램 / 03_과거자료
PROJECT_ROOT = os.path.dirname(os.path.dirname(BASE_DIR))  # "8. 공동기금"
ARCHIVE_ROOT = os.path.join(PROJECT_ROOT, "03_과거자료")
SKIP_DIRS = set()
EXTS = {".pdf", ".hwp", ".hwpx", ".xlsx", ".xls", ".docx", ".doc", ".zip", ".jpg", ".png", ".jpeg", ".tif"}

# 등록 기금 별칭 (별칭 → short_name) — 폴더·파일명에서 기금 식별
ALIASES = {
    "T공동": "T공동 사내", "P사내": "T공동 사내",
    "플러스": "플러스 공동", "케이원": "플러스 공동",
    "이비공동": "이비공동", "EB 공동": "이비공동", "이벌브": "이비공동",
    "가치를만들": "W공동", "가치를 만들": "W공동", "수영로지콘": "W공동",
    "Y공동": "Y공동", "건설Y공동": "Y공동",
    "청신": "X공동", "청원건설": "X공동",
    "배경": "Z공동", "O사내": "O사내", "동경이엔지": "O사내",
    "V공동": "V공동", "○○산업": "○○산업사내", "N사내": "N사내",
    "○○건설": "○○건설사내", "다움": "다움사내",
    "더행복한": "충남 1호",
}
for n in range(1, 13):
    ALIASES[f"충남공동근로복지기금{n}호"] = f"충남 {n}호"
    ALIASES[f"충남공동기금 {n}호"] = f"충남 {n}호"
    ALIASES[f"충남공동기금{n}호"] = f"충남 {n}호"
    ALIASES[f"{n}호기금"] = f"충남 {n}호"
for n in range(1, 5):
    ALIASES[f"경기공동근로복지기금{n}호"] = f"경기 {n}호"
    ALIASES[f"경기지역공동기금{n}호"] = f"경기 {n}호"
ALIASES["더행복한충남공동근로복지기금2호"] = "충남 2호"
ALIASES["더행복한충남공동근로복지기금3호"] = "충남 3호"
ALIASES["더행복한 2호"] = "충남 2호"
ALIASES["더행복한 3호"] = "충남 3호"

# 기록보관 기금 별칭 (등록 완료 → ALIASES에 연결)
for a, s in [("누리공동", "누리공동"), ("누리 공근", "누리공동"), ("엘라이트", "엘라이트"),
             ("지웰삼성", "지웰삼성"), ("정우통상", "정우통상"), ("비츠로밀텍", "비츠로밀텍"),
             ("비츠로셀", "비츠로셀"), ("대흥공동", "대흥공동"), ("하동공동", "하동공동"),
             ("함께보다", "함께보다"), ("U사내", "U사내"), ("엔스코", "엔스코"),
             ("앤스코", "엔스코"), ("신명서진", "신명서진"), ("논산공동", "논산공동"),
             ("논산상생", "논산공동")]:
    ALIASES[a] = s
LEGACY = ["금성침대", "에이치와이", "청양군청"]

DOC_TYPES = [
    ("인가", ["인가증", "인가신청", "노동부인가", "설립인가"]),
    ("정관", ["정관"]),
    ("등기", ["등기부", "등기신청", "법인등기", "등기소", "변경등기", "설립등기"]),
    ("고유번호", ["고유번호"]),
    ("통장·거래", ["통장", "거래내역", "입출금", "이체내역"]),
    ("결산", ["결산"]),
    ("운영상황보고", ["운영상황"]),
    ("사업계획", ["사업계획"]),
    ("지원금", ["지원신청", "지원금", "지원결정", "인센티브"]),
    ("회의록", ["회의록", "회의자료"]),
    ("출연", ["출연확인", "출연약정", "출자자"]),
    ("합의·확인서", ["설립합의", "설립 확인", "설립확인"]),
    ("해산·청산", ["해산", "청산"]),
]


def classify_doc(fname):
    for dtype, kws in DOC_TYPES:
        for kw in kws:
            if kw in fname:
                return dtype
    return None


def find_fund(path_str, alias_sorted):
    for alias, short in alias_sorted:
        if alias in path_str:
            return short
    return None


def main():
    register = "--register" in sys.argv
    init_db()
    con = connect()
    funds = {r["short_name"]: r["fund_id"] for r in
             con.execute("SELECT fund_id, short_name FROM funds").fetchall()}
    # 별칭 정합: short_name이 실제 DB에 있는 것만 (부분일치 허용)
    def resolve(short):
        if short in funds:
            return short
        for k in funds:
            if short in k or k in short:
                return k
        return None

    alias_sorted = sorted(ALIASES.items(), key=lambda x: -len(x[0]))
    matrix = defaultdict(lambda: defaultdict(list))     # fund → dtype → paths
    legacy_hits = defaultdict(int)
    unclassified = 0
    total = 0
    for root, dirs, files in os.walk(ARCHIVE_ROOT):
        rel_root = os.path.relpath(root, ARCHIVE_ROOT)
        top = rel_root.split(os.sep)[0]
        if top in SKIP_DIRS:
            dirs[:] = []
            continue
        for fn in files:
            if fn.startswith("~$") or os.path.splitext(fn)[1].lower() not in EXTS:
                continue
            total += 1
            rel = os.path.join(rel_root, fn)
            fund_short = find_fund(rel, alias_sorted)
            dtype = classify_doc(fn)
            if fund_short:
                r = resolve(fund_short)
                if r and dtype:
                    matrix[r][dtype].append(rel)
                elif r:
                    matrix[r]["기타"].append(rel)
                continue
            for lg in LEGACY:
                if lg in rel:
                    legacy_hits[lg] += 1
                    break
            else:
                unclassified += 1

    # ── 리포트 저장 ──
    scan_dir = os.path.join(BASE_DIR, "scan")
    os.makedirs(scan_dir, exist_ok=True)
    report = {
        "scanned_at": datetime.now().isoformat(timespec="seconds"),
        "total_files": total,
        "funds": {f: {d: len(v) for d, v in dd.items()} for f, dd in matrix.items()},
        "legacy_unregistered": dict(legacy_hits),
        "unclassified": unclassified,
    }
    with open(os.path.join(scan_dir, "scan_report.json"), "w", encoding="utf-8") as f:
        json.dump({**report, "paths": {f: {d: v for d, v in dd.items()}
                                       for f, dd in matrix.items()}},
                  f, ensure_ascii=False, indent=1)

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "보유매트릭스"
    dtypes = [d for d, _ in DOC_TYPES] + ["기타"]
    ws.append(["기금"] + dtypes + ["합계"])
    for short in sorted(matrix.keys()):
        row = [short] + [len(matrix[short].get(d, [])) for d in dtypes]
        ws.append(row + [sum(row[1:])])
    ws2 = wb.create_sheet("미등록기금후보")
    ws2.append(["별칭", "파일수", "처리방안"])
    for lg, n in sorted(legacy_hits.items(), key=lambda x: -x[1]):
        ws2.append([lg, n, "기록보관 기금으로 등록 여부 결정"])
    ws3 = wb.create_sheet("파일목록")
    ws3.append(["기금", "서류종류", "경로"])
    for short, dd in sorted(matrix.items()):
        for d, paths in dd.items():
            for p in paths:
                ws3.append([short, d, p])
    wb.save(os.path.join(scan_dir, "스캔리포트.xlsx"))

    # ── 문서철 등록 (선택) ──
    reg_count = 0
    if register:
        for short, dd in matrix.items():
            fid = funds[short]
            for dtype, paths in dd.items():
                for p in paths:
                    exists = con.execute(
                        "SELECT 1 FROM documents WHERE fund_id=? AND note=?", (fid, p)).fetchone()
                    if exists:
                        continue
                    n = con.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
                    did = f"DOC-{n + 1:06d}"
                    con.execute(
                        "INSERT INTO documents(document_id,fund_id,doc_kind,title,status,note)"
                        " VALUES(?,?,?,?,'archive',?)",
                        (did, fid, dtype, os.path.basename(p), p))
                    reg_count += 1
        audit(con, "documents", "scan", "import", after=f"{reg_count}건 등록")
        con.commit()

    print(f"스캔 {total}개 파일 / 기금 매칭 {sum(len(v) for dd in matrix.values() for v in dd.values())}건"
          f" / 미등록기금 후보 {sum(legacy_hits.values())}건 / 미분류 {unclassified}건")
    if register:
        print(f"문서철 등록: {reg_count}건")
    print(f"리포트: fund-erp/scan/스캔리포트.xlsx")
    con.close()


if __name__ == "__main__":
    main()
