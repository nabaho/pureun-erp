# -*- coding: utf-8 -*-
"""기금출연확인서.xlsm → 실데이터 제거한 빈 템플릿(.xlsx) 생성
   원본은 충남8호 참여사 38곳의 대표자·근로자 성명/생년월일(PII)이 들어있어 그대로 쓰면 안 됨."""
import io, sys, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl

SRC = sys.argv[1]
DST = sys.argv[2]

wb = openpyxl.load_workbook(SRC, data_only=False, keep_vba=False)

# 1) 기업정보: 2행 이하 실데이터 삭제(수식 열 I는 서식 유지 위해 1행만 남김)
ws = wb['기업정보']
cleared = 0
for r in range(2, ws.max_row + 1):
    for c in range(1, 10):
        if ws.cell(r, c).value is not None:
            ws.cell(r, c).value = None; cleared += 1
print("기업정보 지운 셀:", cleared)

# 2) 확인서 본문의 특정 기금명 → 치환 토큰
ws2 = wb['6.기금출연확인서']
subs = 0
for row in ws2.iter_rows():
    for c in row:
        v = c.value
        if isinstance(v, str) and '충남근로복지기금' in v:
            old = v
            v = re.sub(r'충남근로복지기금\s*8?호?', '{{기금명}}', v)
            c.value = v; subs += 1
            print("  %-6s %s → %s" % (c.coordinate, old.strip()[:38], v.strip()[:38]))
# 기금명이 두 셀에 걸쳐 끊긴 잔재 정리 ("8호에 출연할 것을 확인함." → "에 출연할 …")
for row in ws2.iter_rows():
    for c in row:
        if isinstance(c.value, str) and re.match(r'^\s*\d+호에\s', c.value):
            old = c.value; c.value = re.sub(r'^\s*\d+호에', '에', c.value)
            print("  %-6s 잔재정리 %s → %s" % (c.coordinate, old.strip()[:30], c.value.strip()[:30]))
# 연번 초기화
ws2['CJ3'] = 1
# 날짜 비움
for addr in ['AD23']:
    if ws2[addr].value: ws2[addr] = None

wb.save(DST)
print("치환:", subs, "→ 저장:", DST, os.path.getsize(DST), "bytes")

# 검증: 재로드해 실데이터 잔존 확인
chk = openpyxl.load_workbook(DST)
REAL = ['BM금속','변미숙','손항준','충남근로복지기금','그린자동차']
bad = []
for s in chk.worksheets:
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                for w in REAL:
                    if w in c.value: bad.append((s.title, c.coordinate, c.value[:30]))
print("잔존 실데이터:", bad if bad else "없음 ✓")
print("병합셀 보존:", {s.title: len(s.merged_cells.ranges) for s in chk.worksheets})
