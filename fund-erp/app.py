# -*- coding: utf-8 -*-
"""근로복지기금 통합 운영시스템 — FastAPI 백엔드 (계획서 v0.5 / 1단계)
실행: python -m uvicorn app:app --port 8777 --reload
"""
import io, json, os, re
from datetime import datetime
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import openpyxl

from db import connect, init_db, audit, rows_to_dicts, next_id

BASE = os.path.dirname(os.path.abspath(__file__))
app = FastAPI(title="근로복지기금 통합 운영시스템")

init_db()

# ── 사업장 마스터/연도 필드 정의 (인라인 편집 허용 목록) ──
SITE_MASTER_FIELDS = {"name", "biz_no", "corp_no", "ceo", "ceo2", "address", "biz_type",
                      "company_size", "is_primary", "join_date", "leave_date", "note",
                      "seq_label", "contacts"}
SITE_YEAR_FIELDS = {"employees", "capital", "worker_rep", "contribution", "note", "base_date"}
FUND_FIELDS = {"name", "short_name", "fund_type", "region", "status", "inka_no", "inka_date",
               "corp_reg_no", "reg_date", "registry_office", "tax_id_no", "tax_office",
               "address", "phone", "chairman", "labor_office", "note",
               "stg_estab", "stg_reg", "stg_ops", "stg_close"}


class PatchBody(BaseModel):
    field: str
    value: str | int | None = None
    year: int | None = None


class SiteCreate(BaseModel):
    fund_id: str
    name: str
    biz_no: str = ""
    ceo: str = ""
    address: str = ""
    biz_type: str = ""
    year: int | None = None
    employees: int | None = None
    worker_rep: str = ""


class CloseBody(BaseModel):
    closed_date: str
    reason: str
    detail: str = ""


class ContactBody(BaseModel):
    agency_id: str
    name: str
    dept: str = ""
    position: str = ""
    phone: str = ""
    email: str = ""


def norm_bizno(v):
    d = re.sub(r"[^0-9]", "", str(v or ""))[:10]
    return f"{d[:3]}-{d[3:5]}-{d[5:]}" if len(d) == 10 else str(v or "").strip()


# ═══════════ 대시보드 · 기금 ═══════════

@app.get("/api/dashboard")
def dashboard(year: int = datetime.now().year):
    con = connect()
    funds = rows_to_dicts(con.execute(
        "SELECT * FROM funds ORDER BY CASE WHEN mgmt_type='기록보관' THEN 1 ELSE 0 END,"
        " region DESC, CAST(replace(replace(short_name,'충남 ',''),"
        "'호','') AS INTEGER), short_name").fetchall())
    for f in funds:
        fid = f["fund_id"]
        f["site_count"] = con.execute(
            "SELECT COUNT(*) FROM sites WHERE fund_id=? AND status='active'", (fid,)).fetchone()[0]
        f["snap_count"] = con.execute(
            "SELECT COUNT(*) FROM site_histories h JOIN sites s ON s.site_id=h.site_id"
            " WHERE s.fund_id=? AND h.year=?", (fid, year)).fetchone()[0]
        f["snap_locked"] = con.execute(
            "SELECT COUNT(*) FROM site_histories h JOIN sites s ON s.site_id=h.site_id"
            " WHERE s.fund_id=? AND h.year=? AND h.locked=1", (fid, year)).fetchone()[0]
        # 경고: 근로자대표/근로자수 누락 (해당 연도)
        f["warn_missing"] = con.execute(
            "SELECT COUNT(*) FROM site_histories h JOIN sites s ON s.site_id=h.site_id"
            " WHERE s.fund_id=? AND h.year=? AND s.status='active'"
            " AND (h.worker_rep IS NULL OR h.worker_rep='' OR h.employees IS NULL)",
            (fid, year)).fetchone()[0]
        ags = rows_to_dicts(con.execute(
            "SELECT a.name, fa.role FROM fund_agencies fa JOIN agencies a ON a.agency_id=fa.agency_id"
            " WHERE fa.fund_id=? ORDER BY fa.role", (fid,)).fetchall())
        f["supporters"] = [x["name"] for x in ags if x["role"] in ("출연지자체", "지원기관")]
        f["trustee"] = next((x["name"] for x in ags if x["role"] == "수탁기관"), "")
        # 연간 운영 진행 + 운영상황보고 기한 (기록보관 기금 제외)
        agg = con.execute(
            "SELECT COUNT(*) t, COALESCE(SUM(status='done'),0) d FROM tasks WHERE fund_id=? AND year=?",
            (fid, year)).fetchone()
        f["annual_total"], f["annual_done"] = agg["t"], agg["d"]
        rpt = con.execute(
            "SELECT status FROM tasks WHERE fund_id=? AND year=? AND code='RPT-03'",
            (fid, year)).fetchone()
        f["report_due"] = f"{year + 1}-03-31"          # 회계연도 종료 후 3개월
        f["report_done"] = bool(rpt and rpt["status"] == "done")
        # 임원 임기 만료 임박(90일)·만료
        f["role_expiring"] = con.execute(
            "SELECT COUNT(*) FROM fund_roles WHERE fund_id=? AND status='active' AND term_end!=''"
            " AND term_end <= date('now','localtime','+90 day')", (fid,)).fetchone()[0]
    con.close()
    # 보고기한 상태 집계 (기록보관 제외)
    from datetime import date as _date
    today = datetime.now().strftime("%Y-%m-%d")
    op = [f for f in funds if f.get("mgmt_type") not in ("기록보관", "계약종료")]
    overdue = [f["short_name"] for f in op if not f["report_done"] and f["report_due"] < today]
    soon = []
    for f in op:
        if f["report_done"] or f["report_due"] < today:
            continue
        try:
            dd = (_date.fromisoformat(f["report_due"]) - _date.fromisoformat(today)).days
            if dd <= 60:
                soon.append(f["short_name"])
        except ValueError:
            pass
    con2 = connect()
    staff = [r["name"] for r in rows_to_dicts(
        con2.execute("SELECT name FROM staff WHERE active=1 ORDER BY staff_id").fetchall())]
    con2.close()
    return {"year": year, "funds": funds, "staff": staff,
            "report_summary": {"overdue": overdue, "soon": soon,
                               "done": sum(1 for f in op if f["report_done"]),
                               "operating": len(op)}}


@app.post("/api/annual/seed-all")
def seed_all_annual(year: int = datetime.now().year):
    """운영 중 기금 전체에 해당 연도 연간 일정 생성 (대시보드 진행 표시용)"""
    con = connect()
    funds = rows_to_dicts(con.execute(
        "SELECT fund_id, fund_type FROM funds WHERE mgmt_type NOT IN ('기록보관','계약종료') OR mgmt_type IS NULL"
    ).fetchall())
    n = 0
    for f in funds:
        before = con.execute("SELECT COUNT(*) FROM tasks WHERE fund_id=? AND year=?",
                             (f["fund_id"], year)).fetchone()[0]
        _seed_annual(con, f["fund_id"], year, f["fund_type"])
        if con.execute("SELECT COUNT(*) FROM tasks WHERE fund_id=? AND year=?",
                       (f["fund_id"], year)).fetchone()[0] > before:
            n += 1
    con.commit()
    con.close()
    return {"ok": True, "seeded_funds": n}


@app.get("/api/funds/{fund_id}")
def fund_detail(fund_id: str):
    con = connect()
    f = con.execute("SELECT * FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    fund = dict(f)
    fund["agencies"] = rows_to_dicts(con.execute(
        "SELECT fa.link_id, fa.role, fa.contract_start, fa.contract_end, fa.note,"
        " a.agency_id, a.name, a.agency_type, a.phone"
        " FROM fund_agencies fa JOIN agencies a ON a.agency_id=fa.agency_id"
        " WHERE fa.fund_id=? ORDER BY CASE fa.role WHEN '출연지자체' THEN 1 WHEN '지원기관' THEN 2"
        " WHEN '위탁기관' THEN 3 WHEN '수탁기관' THEN 4 ELSE 5 END", (fund_id,)).fetchall())
    for ag in fund["agencies"]:
        ag["contacts"] = rows_to_dicts(con.execute(
            "SELECT * FROM agency_contacts WHERE agency_id=? ORDER BY is_current DESC, contact_id DESC",
            (ag["agency_id"],)).fetchall())
    con.close()
    return fund


@app.patch("/api/funds/{fund_id}")
def patch_fund(fund_id: str, body: PatchBody):
    if body.field not in FUND_FIELDS:
        raise HTTPException(400, f"수정 불가 필드: {body.field}")
    con = connect()
    old = con.execute(f"SELECT {body.field} FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not old:
        con.close()
        raise HTTPException(404, "기금 없음")
    con.execute(f"UPDATE funds SET {body.field}=?, updated_at=datetime('now','localtime')"
                " WHERE fund_id=?", (body.value, fund_id))
    audit(con, "funds", fund_id, "update", body.field, old[0], body.value)
    con.commit()
    con.close()
    return {"ok": True}


# ═══════════ 참여사업장 (pu-erp 업체관리 패턴) ═══════════

@app.get("/api/funds/{fund_id}/sites")
def fund_sites(fund_id: str, year: int = datetime.now().year, status: str = "active"):
    con = connect()
    sites = rows_to_dicts(con.execute(
        "SELECT s.*, h.employees, h.capital, h.worker_rep, h.contribution, h.locked AS snap_locked,"
        " h.base_date, hp.employees AS prev_employees"
        " FROM sites s"
        " LEFT JOIN site_histories h  ON h.site_id=s.site_id AND h.year=?"
        " LEFT JOIN site_histories hp ON hp.site_id=s.site_id AND hp.year=?"
        " WHERE s.fund_id=? AND s.status=?"
        " ORDER BY s.seq_label, s.site_id",
        (year, year - 1, fund_id, status)).fetchall())
    for s in sites:
        s["contacts"] = json.loads(s.get("contacts") or "[]")
        # 급변동 경고 (±50% & 5명 이상)
        pe, ce = s.get("prev_employees"), s.get("employees")
        s["warn_jump"] = bool(pe and ce is not None and pe > 0
                              and abs(ce - pe) >= 5 and abs(ce - pe) / pe >= 0.5)
    counts = {r["status"]: r["n"] for r in rows_to_dicts(con.execute(
        "SELECT status, COUNT(*) n FROM sites WHERE fund_id=? GROUP BY status", (fund_id,)).fetchall())}
    con.close()
    return {"year": year, "sites": sites,
            "counts": {"active": counts.get("active", 0), "closed": counts.get("closed", 0)}}


@app.post("/api/sites")
def create_site(body: SiteCreate):
    con = connect()
    biz = norm_bizno(body.biz_no)
    # 중복 감지 (detectDuplicates 패턴)
    if biz:
        dup = con.execute("SELECT site_id, name, fund_id FROM sites WHERE biz_no=?", (biz,)).fetchone()
        if dup:
            con.close()
            raise HTTPException(409, f"중복 사업자번호: {dup['name']} ({dup['fund_id']})")
    sid = next_id(con, "sites", "site_id", "SITE")
    con.execute(
        "INSERT INTO sites(site_id,fund_id,name,biz_no,ceo,address,biz_type,join_date)"
        " VALUES(?,?,?,?,?,?,?,date('now','localtime'))",
        (sid, body.fund_id, body.name, biz, body.ceo, body.address, body.biz_type))
    if body.year:
        con.execute(
            "INSERT INTO site_histories(site_id,year,employees,worker_rep) VALUES(?,?,?,?)",
            (sid, body.year, body.employees, body.worker_rep))
    audit(con, "sites", sid, "create", after=body.name)
    con.commit()
    con.close()
    return {"ok": True, "site_id": sid}


@app.patch("/api/sites/{site_id}")
def patch_site(site_id: str, body: PatchBody):
    con = connect()
    if body.field in SITE_MASTER_FIELDS:
        old = con.execute(f"SELECT {body.field} FROM sites WHERE site_id=?", (site_id,)).fetchone()
        if not old:
            con.close()
            raise HTTPException(404, "사업장 없음")
        val = body.value
        if body.field == "biz_no":
            val = norm_bizno(val)
        con.execute(f"UPDATE sites SET {body.field}=?, updated_at=datetime('now','localtime')"
                    " WHERE site_id=?", (val, site_id))
        audit(con, "sites", site_id, "update", body.field, old[0], val)
    elif body.field in SITE_YEAR_FIELDS:
        if not body.year:
            con.close()
            raise HTTPException(400, "연도 필요")
        lock = con.execute("SELECT locked FROM site_histories WHERE site_id=? AND year=?",
                           (site_id, body.year)).fetchone()
        if lock and lock[0]:
            con.close()
            raise HTTPException(423, f"{body.year}년 스냅샷 잠금 상태 — 잠금 해제 후 수정")
        val = body.value
        if body.field in ("employees", "capital", "contribution") and val is not None and str(val) != "":
            val = int(re.sub(r"[^0-9-]", "", str(val)) or 0)
        elif body.field in ("employees", "capital", "contribution"):
            val = None
        con.execute(
            f"INSERT INTO site_histories(site_id,year,{body.field}) VALUES(?,?,?)"
            f" ON CONFLICT(site_id,year) DO UPDATE SET {body.field}=excluded.{body.field}",
            (site_id, body.year, val))
        audit(con, "site_histories", f"{site_id}/{body.year}", "update", body.field, "", val)
    else:
        con.close()
        raise HTTPException(400, f"수정 불가 필드: {body.field}")
    con.commit()
    con.close()
    return {"ok": True}


@app.post("/api/sites/{site_id}/close")
def close_site(site_id: str, body: CloseBody):
    """탈퇴·폐업 — 삭제 아님 (pu-erp 종료 모달 패턴)"""
    con = connect()
    row = con.execute("SELECT name FROM sites WHERE site_id=?", (site_id,)).fetchone()
    if not row:
        con.close()
        raise HTTPException(404, "사업장 없음")
    con.execute(
        "UPDATE sites SET status='closed', leave_date=?, closed_reason=?, closed_detail=?,"
        " updated_at=datetime('now','localtime') WHERE site_id=?",
        (body.closed_date, body.reason, body.detail, site_id))
    audit(con, "sites", site_id, "close", "status", "active", f"closed:{body.reason}")
    con.commit()
    con.close()
    return {"ok": True}


@app.post("/api/sites/{site_id}/reopen")
def reopen_site(site_id: str):
    con = connect()
    con.execute("UPDATE sites SET status='active', leave_date='', closed_reason='',"
                " closed_detail='' WHERE site_id=?", (site_id,))
    audit(con, "sites", site_id, "update", "status", "closed", "active")
    con.commit()
    con.close()
    return {"ok": True}


# ═══════════ 연도 스냅샷 잠금 · 엑셀 왕복 ═══════════

@app.post("/api/funds/{fund_id}/snapshot/{year}/lock")
def lock_snapshot(fund_id: str, year: int, unlock: bool = False):
    con = connect()
    v = 0 if unlock else 1
    con.execute(
        "UPDATE site_histories SET locked=? WHERE year=? AND site_id IN"
        " (SELECT site_id FROM sites WHERE fund_id=?)", (v, year, fund_id))
    audit(con, "site_histories", f"{fund_id}/{year}", "lock" if v else "update",
          "locked", "", str(v))
    con.commit()
    con.close()
    return {"ok": True, "locked": bool(v)}


EXPORT_HEADERS = ["연번", "상호", "대표자", "소재지", "상시근로자수", "업종", "근로자대표",
                  "사업자등록번호", "담당자", "직무", "전화번호", "휴대폰번호", "이메일"]


@app.get("/api/funds/{fund_id}/sites/export")
def export_sites(fund_id: str, year: int = datetime.now().year):
    """갱신용 엑셀 내보내기 — 전년(현재) 값 채워서 제공, 연번은 텍스트 서식"""
    con = connect()
    data = fund_sites(fund_id, year)["sites"]
    fund = con.execute("SELECT short_name FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    con.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}갱신"
    ws.append(EXPORT_HEADERS)
    for s in data:
        c = (s["contacts"] or [{}])[0] if s["contacts"] else {}
        cell = ws.append([
            s.get("seq_label") or "", s["name"], s.get("ceo") or "", s.get("address") or "",
            s.get("employees"), s.get("biz_type") or "", s.get("worker_rep") or "",
            s.get("biz_no") or "", c.get("name", ""), c.get("position", ""),
            c.get("phone", ""), c.get("mobile", ""), c.get("email", ""),
        ])
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].number_format = "@"  # 연번 날짜 변환 방지
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = f"{(fund['short_name'] if fund else fund_id).replace(' ', '')}_{year}_갱신용.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


@app.post("/api/funds/{fund_id}/sites/import")
async def import_sites(fund_id: str, year: int, file: UploadFile = File(...)):
    """갱신 엑셀 가져오기 — 사업자번호 매칭, 변경분만 upsert, 검증 리포트 반환"""
    raw = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows[:10])
               if r and any("상호" in str(x or "") or "사업장" in str(x or "") for x in r)), None)
    if hi is None:
        raise HTTPException(400, "헤더 행(상호/사업장)을 찾을 수 없음")
    hdr = [str(x or "").strip() for x in rows[hi]]

    def col(*names):
        for i, hcol in enumerate(hdr):
            if any(n in hcol for n in names):
                return i
        return -1

    cName, cBiz = col("상호", "사업장명", "업체"), col("사업자")
    cEmp, cRep = col("상시근로자", "근로자수"), col("근로자대표")
    cCeo, cAddr = col("대표자"), col("소재지", "주소")
    if cName < 0:
        raise HTTPException(400, "상호 컬럼 없음")

    con = connect()
    existing = rows_to_dicts(con.execute(
        "SELECT site_id, name, biz_no FROM sites WHERE fund_id=?", (fund_id,)).fetchall())
    by_biz = {re.sub(r"[^0-9]", "", s["biz_no"] or ""): s for s in existing if s["biz_no"]}
    by_name = {re.sub(r"\s|\(주\)|㈜|주식회사", "", s["name"]): s for s in existing}
    matched = created = 0
    jumps, missing, unmatched = [], [], []
    for r in rows[hi + 1:]:
        if not r or all(x is None or str(x).strip() == "" for x in r):
            continue
        name = str(r[cName] or "").strip()
        if not name:
            continue
        biz_d = re.sub(r"[^0-9]", "", str(r[cBiz] or "")) if cBiz >= 0 else ""
        site = by_biz.get(biz_d) or by_name.get(re.sub(r"\s|\(주\)|㈜|주식회사", "", name))
        emp = None
        if cEmp >= 0 and r[cEmp] is not None and str(r[cEmp]).strip() != "":
            try:
                emp = int(float(str(r[cEmp]).replace(",", "")))
            except ValueError:
                emp = None
        wrep = str(r[cRep] or "").strip() if cRep >= 0 else ""
        if not site:
            # 신규 사업장 추가
            sid = next_id(con, "sites", "site_id", "SITE")
            con.execute(
                "INSERT INTO sites(site_id,fund_id,name,biz_no,ceo,address,join_date)"
                " VALUES(?,?,?,?,?,?,date('now','localtime'))",
                (sid, fund_id, name, norm_bizno(biz_d),
                 str(r[cCeo] or "").strip() if cCeo >= 0 else "",
                 str(r[cAddr] or "").strip() if cAddr >= 0 else ""))
            audit(con, "sites", sid, "create", after=f"{name} (엑셀 가져오기)")
            created += 1
            unmatched.append(name)
        else:
            sid = site["site_id"]
            matched += 1
        prev = con.execute("SELECT employees FROM site_histories WHERE site_id=? AND year=?",
                           (sid, year - 1)).fetchone()
        lock = con.execute("SELECT locked FROM site_histories WHERE site_id=? AND year=?",
                           (sid, year)).fetchone()
        if lock and lock[0]:
            continue  # 잠긴 스냅샷은 건너뜀
        con.execute(
            "INSERT INTO site_histories(site_id,year,employees,worker_rep) VALUES(?,?,?,?)"
            " ON CONFLICT(site_id,year) DO UPDATE SET employees=excluded.employees,"
            " worker_rep=excluded.worker_rep", (sid, year, emp, wrep))
        if emp is None or not wrep:
            missing.append(name)
        if prev and prev[0] and emp is not None and prev[0] > 0 \
                and abs(emp - prev[0]) >= 5 and abs(emp - prev[0]) / prev[0] >= 0.5:
            jumps.append(f"{name} {prev[0]}→{emp}")
    audit(con, "sites", fund_id, "import", after=f"{file.filename} 매칭{matched} 신규{created}")
    con.commit()
    con.close()
    return {"ok": True, "matched": matched, "created": created,
            "warn_jumps": jumps, "warn_missing": missing[:50], "new_names": unmatched[:50]}


# ═══════════ 관계기관 · 담당자 ═══════════

@app.get("/api/agencies")
def agencies():
    con = connect()
    out = rows_to_dicts(con.execute("SELECT * FROM agencies ORDER BY agency_type, name").fetchall())
    con.close()
    return out


@app.post("/api/agency_contacts")
def add_contact(body: ContactBody):
    con = connect()
    # 기존 현재 담당자 → 이력 보존 (교체)
    con.execute("UPDATE agency_contacts SET is_current=0, end_date=date('now','localtime')"
                " WHERE agency_id=? AND is_current=1 AND name<>?", (body.agency_id, body.name))
    con.execute(
        "INSERT INTO agency_contacts(agency_id,name,dept,position,phone,email,start_date)"
        " VALUES(?,?,?,?,?,?,date('now','localtime'))",
        (body.agency_id, body.name, body.dept, body.position, body.phone, body.email))
    audit(con, "agency_contacts", body.agency_id, "create", after=body.name)
    con.commit()
    con.close()
    return {"ok": True}


# ═══════════ 문서철 · 일괄수정 · 서류 생성 (2단계) ═══════════
import docgen as _docgen
from fastapi.responses import HTMLResponse

ARCHIVE_DIR = os.path.join(os.path.dirname(os.path.dirname(BASE)), "03_과거자료")


@app.get("/api/funds/{fund_id}/documents")
def fund_documents(fund_id: str, kind: str = ""):
    con = connect()
    q = "SELECT document_id, doc_kind, title, status, note FROM documents WHERE fund_id=?"
    args = [fund_id]
    if kind:
        q += " AND doc_kind=?"
        args.append(kind)
    docs = rows_to_dicts(con.execute(q + " ORDER BY doc_kind, title", args).fetchall())
    counts = rows_to_dicts(con.execute(
        "SELECT doc_kind, COUNT(*) n FROM documents WHERE fund_id=? GROUP BY doc_kind"
        " ORDER BY n DESC", (fund_id,)).fetchall())
    con.close()
    return {"docs": docs, "counts": counts}


UPLOADS_DIR = os.path.join(BASE, "uploads")
SCAN_KINDS = ["설립인가", "법인등기", "변경등기", "정관변경", "고유번호", "지원금",
              "회의록", "출연·기본재산", "결산", "운영상황보고", "해산·청산", "기타"]


@app.get("/api/docfile/{document_id}")
def doc_file(document_id: str):
    con = connect()
    row = con.execute("SELECT note, title, status, file_path FROM documents WHERE document_id=?",
                      (document_id,)).fetchone()
    con.close()
    if not row:
        raise HTTPException(404, "문서 없음")
    # 업로드 스캔본은 uploads 폴더, 과거자료는 03_과거자료
    if row["status"] == "upload" and row["file_path"]:
        path = os.path.normpath(os.path.join(UPLOADS_DIR, row["file_path"]))
        base = os.path.normpath(UPLOADS_DIR)
    else:
        path = os.path.normpath(os.path.join(ARCHIVE_DIR, row["note"]))
        base = os.path.normpath(ARCHIVE_DIR)
    if not path.startswith(base):
        raise HTTPException(403, "잘못된 경로")
    if not os.path.exists(path):
        raise HTTPException(404, f"파일 없음: {row['title']}")
    return FileResponse(path, filename=os.path.basename(path))


@app.post("/api/funds/{fund_id}/documents/upload")
async def upload_document(fund_id: str, doc_kind: str = "기타", doc_date: str = "",
                          note: str = "", file: UploadFile = File(...)):
    con = connect()
    if not con.execute("SELECT 1 FROM funds WHERE fund_id=?", (fund_id,)).fetchone():
        con.close()
        raise HTTPException(404, "기금 없음")
    fdir = os.path.join(UPLOADS_DIR, fund_id)
    os.makedirs(fdir, exist_ok=True)
    # 안전한 저장 파일명: 순번 + 원본명
    raw = await file.read()
    n = con.execute("SELECT COUNT(*) FROM documents WHERE fund_id=? AND status='upload'",
                    (fund_id,)).fetchone()[0]
    safe = re.sub(r"[^\w가-힣.\-]", "_", file.filename or "scan")
    stored = f"{n + 1:03d}_{safe}"
    with open(os.path.join(fdir, stored), "wb") as w:
        w.write(raw)
    did = next_id(con, "documents", "document_id", "DOC")
    rel = f"{fund_id}/{stored}"
    con.execute("INSERT INTO documents(document_id,fund_id,doc_kind,title,status,doc_date,note,"
                "file_path,uploaded_at) VALUES(?,?,?,?, 'upload',?,?,?,datetime('now','localtime'))",
                (did, fund_id, doc_kind, file.filename or stored, doc_date, note, rel))
    audit(con, "documents", did, "create", after=f"스캔 업로드 {doc_kind} {file.filename}")
    con.commit()
    con.close()
    return {"ok": True, "document_id": did, "size": len(raw)}


@app.delete("/api/documents/{document_id}")
def delete_upload(document_id: str):
    con = connect()
    row = con.execute("SELECT status, file_path FROM documents WHERE document_id=?",
                      (document_id,)).fetchone()
    if not row:
        con.close()
        raise HTTPException(404, "문서 없음")
    if row["status"] != "upload":
        con.close()
        raise HTTPException(400, "업로드 문서만 삭제할 수 있습니다 (과거자료는 원본 유지)")
    if row["file_path"]:
        p = os.path.join(UPLOADS_DIR, row["file_path"])
        if os.path.exists(p):
            os.remove(p)
    con.execute("DELETE FROM documents WHERE document_id=?", (document_id,))
    audit(con, "documents", document_id, "update", "delete", "", "스캔 삭제")
    con.commit()
    con.close()
    return {"ok": True}


@app.get("/api/funds/{fund_id}/uploads")
def fund_uploads(fund_id: str):
    con = connect()
    rows = rows_to_dicts(con.execute(
        "SELECT document_id, doc_kind, title, doc_date, uploaded_at FROM documents"
        " WHERE fund_id=? AND status='upload' ORDER BY uploaded_at DESC", (fund_id,)).fetchall())
    con.close()
    return {"uploads": rows, "kinds": SCAN_KINDS}


class ProfileBody(BaseModel):
    profile: dict


@app.put("/api/funds/{fund_id}/profile")
def save_profile(fund_id: str, body: ProfileBody):
    con = connect()
    if not con.execute("SELECT 1 FROM funds WHERE fund_id=?", (fund_id,)).fetchone():
        con.close()
        raise HTTPException(404, "기금 없음")
    con.execute("UPDATE funds SET profile=?, updated_at=datetime('now','localtime') WHERE fund_id=?",
                (json.dumps(body.profile, ensure_ascii=False), fund_id))
    audit(con, "funds", fund_id, "update", "profile", "", "일괄수정 저장")
    con.commit()
    con.close()
    return {"ok": True}


@app.get("/api/funds/{fund_id}/docgen/{kind}", response_class=HTMLResponse)
def generate_doc(fund_id: str, kind: str, year: int = datetime.now().year, edited: int = 1):
    if kind not in _docgen.KINDS:
        raise HTTPException(404, "알 수 없는 서류")
    con = connect()
    # 편집 저장본이 있으면 우선 반환 (edited=0으로 원본 강제 재생성 가능)
    if edited:
        e = con.execute("SELECT html FROM doc_edits WHERE fund_id=? AND year=? AND kind=?",
                        (fund_id, year, kind)).fetchone()
        if e and e["html"]:
            con.close()
            return e["html"]
    f = con.execute("SELECT * FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    fund = dict(f)
    sites = rows_to_dicts(con.execute(
        "SELECT s.*, h.employees, h.contribution FROM sites s"
        " LEFT JOIN site_histories h ON h.site_id=s.site_id AND h.year=?"
        " WHERE s.fund_id=? AND s.status='active' ORDER BY s.seq_label", (year, fund_id)).fetchall())
    # 임원 명부(M4)가 있으면 서류 임원 표시에 우선 사용
    roles = rows_to_dicts(con.execute(
        "SELECT person_name, role FROM fund_roles WHERE fund_id=? AND status='active'", (fund_id,)).fetchall())
    if roles:
        prof = json.loads(fund.get("profile") or "{}")
        prof["officers"] = [{"role": r["role"], "name": r["person_name"]} for r in roles]
        fund["profile"] = json.dumps(prof, ensure_ascii=False)
    fs = None
    if kind in ("form15", "settlement"):
        cl = closing(fund_id, year)
        fs = dict(cl.get("financials") or {})
        if kind == "form15":
            fs["welfare"] = welfare_form15(con, fund_id, year)
        if kind == "settlement":
            fs["trial_balance"] = cl.get("trial_balance") or {}
    html = _docgen.render(kind, fund, sites, fs, year)
    # 문서·버전 등록 (초안)
    title = _docgen.KINDS[kind]
    row = con.execute("SELECT document_id FROM documents WHERE fund_id=? AND doc_kind='생성서류'"
                      " AND title=?", (fund_id, title)).fetchone()
    if row:
        did = row[0]
    else:
        n = con.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
        did = f"DOC-{n + 1:06d}"
        con.execute("INSERT INTO documents(document_id,fund_id,doc_kind,title,status,note)"
                    " VALUES(?,?,'생성서류',?,'draft','docgen')", (did, fund_id, title))
    ver = (con.execute("SELECT COALESCE(MAX(ver),0) FROM document_versions WHERE document_id=?",
                       (did,)).fetchone()[0] or 0) + 1
    con.execute("INSERT INTO document_versions(version_id,document_id,ver,snapshot_json)"
                " VALUES(?,?,?,?)",
                (f"{did}-V{ver:02d}", did, ver,
                 json.dumps({"kind": kind, "at": datetime.now().isoformat(timespec='seconds')},
                            ensure_ascii=False)))
    audit(con, "documents", did, "create", "docgen", "", f"{title} V{ver}")
    con.commit()
    con.close()
    return html


# ═══════════ 신규 기금 설립 · 담당자(직원) ═══════════
EST_TEMPLATE = [
    ("EST-01", "상담·유형 판정(사내/공동)·수임 계약", False),
    ("EST-02", "참여사업장 모집·확정 (참여사업장 탭 등록)", True),      # 공동 전용
    ("EST-03", "설립준비위원회 구성 — 노사 동수 확인", False),
    ("EST-04", "출연 계획·재산목록 확정 (일괄수정 ①)", False),
    ("EST-05", "정관 작성 (서류 생성 → 정관)", False),
    ("EST-06", "설립준비위 회의 의결·날인 (회의록)", False),
    ("EST-07", "설립인가신청서(별지 제7호) 제출", False),
    ("EST-08", "인가증 수령 → 인가번호·인가일 등록 (일괄수정 ②)", False),
    ("EST-09", "설립등기 — 인가 후 3주 이내 (등기 8종)", False),
    ("EST-10", "법인등록번호·고유번호 발급 등록", False),
    ("EST-11", "법인 계좌 개설·인감 관리 (회계·결산 탭 계좌)", False),
    ("EST-12", "최초 출연금 입금 확인 → 운영 전환", False),
]
EST_STATUSES = ("설립준비", "인가신청", "등기진행")


class ContractBody(BaseModel):
    action: str                 # end | resume
    date: str = ""
    reason: str = ""


@app.post("/api/funds/{fund_id}/contract")
def fund_contract(fund_id: str, body: ContractBody):
    """계약 종료/재개 — 데이터는 전부 보존, 홈에서 접힌 섹션으로 이동만"""
    con = connect()
    f = con.execute("SELECT mgmt_type, mgmt_prev FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    if body.action == "end":
        con.execute("UPDATE funds SET mgmt_prev=?, mgmt_type='계약종료', contract_end=?,"
                    " contract_end_reason=?, updated_at=datetime('now','localtime') WHERE fund_id=?",
                    (f["mgmt_type"] or "수임", body.date, body.reason, fund_id))
        audit(con, "funds", fund_id, "update", "계약", f["mgmt_type"], f"계약종료 {body.date} {body.reason}")
    elif body.action == "resume":
        con.execute("UPDATE funds SET mgmt_type=?, contract_end='', contract_end_reason='',"
                    " updated_at=datetime('now','localtime') WHERE fund_id=?",
                    (f["mgmt_prev"] or "수임", fund_id))
        audit(con, "funds", fund_id, "update", "계약", "계약종료", f"재개({f['mgmt_prev'] or '수임'})")
    else:
        con.close()
        raise HTTPException(400, "action은 end|resume")
    con.commit()
    con.close()
    return {"ok": True}


@app.get("/api/staff")
def list_staff(all: bool = False):
    """담당자(=푸른이알피 근로자) 목록 + 주/부담당 기금수 + 연결 상태."""
    con = connect()
    where = "" if all else " WHERE s.active=1"
    out = rows_to_dicts(con.execute(
        "SELECT s.staff_id, s.name, s.puerp_uid, s.active, s.note,"
        " (SELECT COUNT(*) FROM fund_staff fs WHERE fs.staff_id=s.staff_id"
        "  AND fs.is_current=1 AND fs.role='정') AS main_cnt,"
        " (SELECT COUNT(*) FROM fund_staff fs WHERE fs.staff_id=s.staff_id"
        "  AND fs.is_current=1 AND fs.role='부') AS sub_cnt"
        f" FROM staff s{where} ORDER BY s.staff_id").fetchall())
    for r in out:
        r["linked"] = bool((r.get("puerp_uid") or "").strip())
    con.close()
    return out


class StaffBody(BaseModel):
    name: str
    puerp_uid: str = ""
    active: int = 1


@app.post("/api/staff")
def create_staff(body: StaffBody):
    """담당자 추가 (이름 중복이면 puerp_uid·active 갱신 = 푸른이알피 연결 설정)."""
    nm = body.name.strip()
    if not nm:
        raise HTTPException(400, "이름을 입력하세요")
    con = connect()
    row = con.execute("SELECT staff_id FROM staff WHERE name=?", (nm,)).fetchone()
    if row:
        sid = row[0]
        con.execute("UPDATE staff SET puerp_uid=?, active=? WHERE staff_id=?",
                    (body.puerp_uid.strip(), 1 if body.active else 0, sid))
    else:
        n = con.execute("SELECT COUNT(*) FROM staff").fetchone()[0]
        sid = f"ST-{n + 1:03d}"
        con.execute("INSERT INTO staff(staff_id,name,puerp_uid,active) VALUES(?,?,?,?)",
                    (sid, nm, body.puerp_uid.strip(), 1 if body.active else 0))
    con.commit()
    con.close()
    return {"ok": True, "staff_id": sid}


@app.patch("/api/staff/{staff_id}")
def update_staff(staff_id: str, body: StaffBody):
    con = connect()
    if not con.execute("SELECT 1 FROM staff WHERE staff_id=?", (staff_id,)).fetchone():
        con.close()
        raise HTTPException(404, "담당자 없음")
    con.execute("UPDATE staff SET name=?, puerp_uid=?, active=? WHERE staff_id=?",
                (body.name.strip(), body.puerp_uid.strip(), 1 if body.active else 0, staff_id))
    con.commit()
    con.close()
    return {"ok": True}


class NewFundBody(BaseModel):
    name: str
    fund_type: str = "공동"
    region: str = ""
    manager_main: str = ""     # staff name
    manager_sub: str = ""
    entry_mode: str = "new"    # new=신규(설립부터) | existing=기존(운영부터)


def _set_fund_staff(con, fund_id, main_name, sub_name, uids=None):
    """주(정)/부 담당 지정 — 기존 담당은 이력으로 종료.
    uids={이름: 푸른이알피 sid} 를 주면 담당자 마스터에 연결키(puerp_uid)까지 심는다."""
    uids = uids or {}
    con.execute("UPDATE fund_staff SET is_current=0, end_date=date('now','localtime')"
                " WHERE fund_id=? AND is_current=1", (fund_id,))
    names = []
    for nm, role in ((main_name, "정"), (sub_name, "부")):
        nm = (nm or "").strip()
        if not nm:
            continue
        uid = str(uids.get(nm, "") or "").strip()
        row = con.execute("SELECT staff_id, puerp_uid FROM staff WHERE name=?", (nm,)).fetchone()
        if not row:
            n = con.execute("SELECT COUNT(*) FROM staff").fetchone()[0]
            sid = f"ST-{n + 1:03d}"
            con.execute("INSERT INTO staff(staff_id,name,puerp_uid) VALUES(?,?,?)", (sid, nm, uid))
        else:
            sid = row[0]
            if uid and not (row[1] or "").strip():   # 기존 연결이 없을 때만 채움
                con.execute("UPDATE staff SET puerp_uid=? WHERE staff_id=?", (uid, sid))
        con.execute("INSERT INTO fund_staff(fund_id,staff_id,role,start_date)"
                    " VALUES(?,?,?,date('now','localtime'))", (fund_id, sid, role))
        names.append(nm + ("" if role == "정" else "(부)"))
    con.execute("UPDATE funds SET manager=? WHERE fund_id=?", (", ".join(names), fund_id))


@app.post("/api/funds")
def create_fund(body: NewFundBody):
    if not body.name.strip():
        raise HTTPException(400, "기금명을 입력하세요")
    con = connect()
    if con.execute("SELECT 1 FROM funds WHERE name=?", (body.name.strip(),)).fetchone():
        con.close()
        raise HTTPException(409, "같은 이름의 기금이 이미 있습니다")
    existing = body.entry_mode == "existing"
    fid = next_id(con, "funds", "fund_id", "FUND")
    if existing:
        # 기존 기금 — 이미 설립·등기 완료, 곧장 운영 관리
        status, stg_estab, stg_reg, stg_ops = "운영", "완료", "완료", "진행"
        after_note = f"{body.name} (기존·운영부터)"
    else:
        # 신규 기금 — 설립부터 서식 만들며 진행
        status, stg_estab, stg_reg, stg_ops = "설립준비", "진행", "없음", "없음"
        after_note = f"{body.name} (신규 설립)"
    con.execute(
        "INSERT INTO funds(fund_id,name,short_name,fund_type,region,status,mgmt_type,"
        "stg_estab,stg_reg,stg_ops,stg_close) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        (fid, body.name.strip(), body.name.strip()[:12], body.fund_type, body.region,
         status, "수임", stg_estab, stg_reg, stg_ops, "없음"))
    audit(con, "funds", fid, "create", after=after_note)
    # 수탁기관·지원기관 자동 연결
    pnp = ensure_agency_app(con, "푸른노무법인", "수탁법인")
    con.execute("INSERT OR IGNORE INTO fund_agencies(fund_id,agency_id,role) VALUES(?,?,'수탁기관')",
                (fid, pnp))
    if body.fund_type == "공동":
        kc = ensure_agency_app(con, "근로복지공단", "공단")
        con.execute("INSERT OR IGNORE INTO fund_agencies(fund_id,agency_id,role) VALUES(?,?,'지원기관')",
                    (fid, kc))
    _set_fund_staff(con, fid, body.manager_main, body.manager_sub)
    # 신규만 설립 체크리스트 seed (year=0). 기존은 설립 단계를 밟지 않음.
    if not existing:
        for i, (code, title, gong_only) in enumerate(EST_TEMPLATE):
            if gong_only and body.fund_type != "공동":
                continue
            con.execute("INSERT INTO tasks(task_id,fund_id,year,code,period,title,status)"
                        " VALUES(?,?,0,?,?,?, 'todo')",
                        (f"TASK-{fid}-EST-{i:02d}-{code}", fid, code, f"{i + 1}단계", title))
    con.commit()
    con.close()
    return {"ok": True, "fund_id": fid, "entry_mode": body.entry_mode}


def ensure_agency_app(con, name, atype):
    row = con.execute("SELECT agency_id FROM agencies WHERE name=?", (name,)).fetchone()
    if row:
        return row[0]
    n = con.execute("SELECT COUNT(*) FROM agencies").fetchone()[0]
    aid = f"AG-{n + 1:04d}"
    con.execute("INSERT INTO agencies(agency_id,name,agency_type) VALUES(?,?,?)", (aid, name, atype))
    return aid


@app.get("/api/funds/{fund_id}/est")
def est_checklist(fund_id: str):
    con = connect()
    f = con.execute("SELECT fund_type, status FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    # seed가 없으면 생성 (기존 이관 기금이 설립 상태로 바뀐 경우 대비)
    if not con.execute("SELECT 1 FROM tasks WHERE fund_id=? AND year=0", (fund_id,)).fetchone():
        for i, (code, title, gong_only) in enumerate(EST_TEMPLATE):
            if gong_only and f["fund_type"] != "공동":
                continue
            con.execute("INSERT INTO tasks(task_id,fund_id,year,code,period,title,status)"
                        " VALUES(?,?,0,?,?,?, 'todo')",
                        (f"TASK-{fund_id}-EST-{i:02d}-{code}", fund_id, code, f"{i + 1}단계", title))
        con.commit()
    tasks = rows_to_dicts(con.execute(
        "SELECT task_id, code, period, title, status, done_date FROM tasks"
        " WHERE fund_id=? AND year=0 ORDER BY task_id", (fund_id,)).fetchall())
    con.close()
    return {"status": f["status"], "tasks": tasks,
            "done": sum(1 for t in tasks if t["status"] == "done"), "total": len(tasks)}


class StaffAssign(BaseModel):
    manager_main: str = ""
    manager_sub: str = ""


@app.post("/api/funds/{fund_id}/staff")
def assign_staff(fund_id: str, body: StaffAssign):
    con = connect()
    if not con.execute("SELECT 1 FROM funds WHERE fund_id=?", (fund_id,)).fetchone():
        con.close()
        raise HTTPException(404, "기금 없음")
    _set_fund_staff(con, fund_id, body.manager_main, body.manager_sub)
    audit(con, "fund_staff", fund_id, "update", "담당",
          "", f"정:{body.manager_main} 부:{body.manager_sub}")
    con.commit()
    con.close()
    return {"ok": True}


# ═══════════ 기금 대장 엑셀 내보내기 ═══════════
@app.get("/api/ledger/export")
def ledger_export(year: int = datetime.now().year):
    con = connect()
    funds = rows_to_dicts(con.execute(
        "SELECT name, fund_type, region, inka_no, inka_date, corp_reg_no, reg_date,"
        " tax_id_no, registry_office, tax_office, labor_office, address, chairman"
        " FROM funds WHERE mgmt_type NOT IN ('기록보관','계약종료') OR mgmt_type IS NULL"
        " ORDER BY fund_type, region DESC, short_name").fetchall())
    con.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기금대장"
    hdr = ["기금명", "유형", "지역", "인가번호", "인가일", "법인등록번호", "설립등기일",
           "고유번호", "관할등기소", "관할세무서", "관할노동청", "이사장", "소재지"]
    ws.append(hdr)
    for f in funds:
        ws.append([f["name"], f["fund_type"], f["region"], f["inka_no"], f["inka_date"],
                   f["corp_reg_no"], f["reg_date"], f["tax_id_no"], f["registry_office"],
                   f["tax_office"], f["labor_office"], f["chairman"], f["address"]])
    for i, w in enumerate([26, 6, 6, 16, 12, 16, 12, 14, 14, 14, 14, 8, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = f"기금대장_{year}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


# ═══════════ 오늘 할 일 요약 (3안 — 홈 최상단) ═══════════
@app.get("/api/todo")
def todo(year: int = datetime.now().year):
    con = connect()
    today = datetime.now().strftime("%Y-%m-%d")
    items = []
    # 1) 운영상황보고 기한 지남 (미제출)
    n = con.execute(
        "SELECT COUNT(*) FROM funds f WHERE (f.mgmt_type NOT IN ('기록보관','계약종료') OR f.mgmt_type IS NULL)"
        " AND NOT EXISTS(SELECT 1 FROM tasks t WHERE t.fund_id=f.fund_id AND t.year=? AND t.code='RPT-03'"
        " AND t.status='done')", (year,)).fetchone()[0]
    if n:
        items.append({"icon": "📄", "level": "bad",
                      "text": f"{year}년 운영상황보고 미제출 {n}건", "go": "board:report"})
    # 2) 사업장 정보 누락 (해당 연도)
    n = con.execute(
        "SELECT COUNT(DISTINCT s.fund_id) FROM site_histories h JOIN sites s ON s.site_id=h.site_id"
        " WHERE h.year=? AND s.status='active' AND (h.worker_rep IS NULL OR h.worker_rep=''"
        " OR h.employees IS NULL)", (year,)).fetchone()[0]
    if n:
        items.append({"icon": "🏢", "level": "warn",
                      "text": f"참여사업장 정보 누락 {n}개 기금", "go": "board:renewal"})
    # 3) 청구 미입금
    row = con.execute(
        "SELECT COUNT(*), COALESCE(SUM(amount),0) FROM billings WHERE year=?"
        " AND invoice_date!='' AND (paid_date='' OR paid_date IS NULL)", (year,)).fetchone()
    if row[0]:
        items.append({"icon": "💰", "level": "warn",
                      "text": f"청구 미입금 {row[0]}건 · {row[1]:,}원", "go": "billing"})
    # 4) 미청구 (계산서 미발행)
    n = con.execute("SELECT COUNT(*) FROM billings WHERE year=? AND (invoice_date='' OR invoice_date IS NULL)",
                    (year,)).fetchone()[0]
    if n:
        items.append({"icon": "🧾", "level": "norm",
                      "text": f"미청구(계산서 미발행) {n}건", "go": "billing"})
    # 5) 임원 임기 만료 임박
    n = con.execute(
        "SELECT COUNT(*) FROM fund_roles WHERE status='active' AND term_end!=''"
        " AND term_end <= date('now','localtime','+90 day')").fetchone()[0]
    if n:
        items.append({"icon": "👔", "level": "norm",
                      "text": f"임원 임기 만료 임박 {n}명", "go": ""})
    # 6) 설립 진행 중
    n = con.execute("SELECT COUNT(*) FROM funds WHERE status IN ('설립준비','인가신청','등기진행')").fetchone()[0]
    if n:
        items.append({"icon": "🏗", "level": "acc", "text": f"설립 진행 중 {n}건", "go": ""})
    con.close()
    return {"date": today, "items": items}


# ═══════════ 청구 대장 (결산 보수 → 푸른이알피 매출 연결) ═══════════
def _ensure_billing(con, fund_id, year, item, note=""):
    con.execute("INSERT OR IGNORE INTO billings(fund_id,year,item,note) VALUES(?,?,?,?)",
                (fund_id, year, item, note))


@app.get("/api/billings")
def list_billings(year: int = datetime.now().year):
    con = connect()
    rows = rows_to_dicts(con.execute(
        "SELECT b.*, f.name fund_name, f.short_name, f.manager, f.fund_type, f.mgmt_type"
        " FROM billings b JOIN funds f ON f.fund_id=b.fund_id WHERE b.year=?"
        " ORDER BY f.fund_type, f.region DESC, f.short_name", (year,)).fetchall())
    con.close()
    for r in rows:
        r["status"] = "입금완료" if r["paid_date"] else ("청구" if r["invoice_date"] else "미청구")
    total = sum(r["amount"] or 0 for r in rows)
    paid = sum(r["amount"] or 0 for r in rows if r["paid_date"])
    return {"year": year, "rows": rows, "total": total, "paid": paid}


class BillingBody(BaseModel):
    fund_id: str
    year: int
    item: str = "결산·운영상황보고 보수"
    amount: int = 0


@app.post("/api/billings")
def add_billing(body: BillingBody):
    con = connect()
    _ensure_billing(con, body.fund_id, body.year, body.item)
    if body.amount:
        con.execute("UPDATE billings SET amount=? WHERE fund_id=? AND year=? AND item=?",
                    (body.amount, body.fund_id, body.year, body.item))
    audit(con, "billings", body.fund_id, "create", after=f"{body.year} {body.item}")
    con.commit()
    con.close()
    return {"ok": True}


class BillingPatch(BaseModel):
    field: str
    value: str | int | None = None


@app.patch("/api/billings/{billing_id}")
def patch_billing(billing_id: int, body: BillingPatch):
    if body.field not in ("amount", "invoice_date", "paid_date", "note", "item", "vat_separate"):
        raise HTTPException(400, "수정 불가 필드")
    v = body.value
    if body.field == "amount":
        v = int(re.sub(r"[^0-9]", "", str(v or "")) or 0)
    con = connect()
    con.execute(f"UPDATE billings SET {body.field}=? WHERE billing_id=?", (v, billing_id))
    audit(con, "billings", str(billing_id), "update", body.field, "", v)
    con.commit()
    con.close()
    return {"ok": True}


@app.delete("/api/billings/{billing_id}")
def del_billing(billing_id: int):
    con = connect()
    con.execute("DELETE FROM billings WHERE billing_id=?", (billing_id,))
    audit(con, "billings", str(billing_id), "update", "delete", "", "삭제")
    con.commit()
    con.close()
    return {"ok": True}


@app.get("/api/billings/export")
def export_billings(year: int = datetime.now().year):
    """푸른이알피 매출(finance_income) 형식 JSON — pu-erp 가져오기 스니펫으로 등록"""
    data = list_billings(year)
    out = []
    for r in data["rows"]:
        out.append({
            "kind": "기금보수", "companyName": r["fund_name"],
            "label": f"{year}년 {r['item']}", "amount": r["amount"] or 0,
            "vatIncluded": not bool(r["vat_separate"]),
            "invoiceDate": r["invoice_date"], "paidDate": r["paid_date"],
            "paid": bool(r["paid_date"]), "year": year,
            "manager": r["manager"], "memo": r["note"] or "",
            "source": "fund-erp",
        })
    body = json.dumps({"fund_billings": out, "exported_at":
                       datetime.now().isoformat(timespec="seconds")}, ensure_ascii=False, indent=1)
    from urllib.parse import quote
    fname = f"기금청구_{year}_푸른이알피용.json"
    return StreamingResponse(io.BytesIO(body.encode("utf-8")), media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


# ═══════════ 서식 자료실 — 공식 법정서식 원본 제공 (00_표준서식 연결) ═══════════
FORMS_DIR = os.path.join(os.path.dirname(os.path.dirname(BASE)), "00_표준서식", "01_공식법정서식_현행")


@app.get("/api/forms")
def list_forms():
    con = connect()
    rows = rows_to_dicts(con.execute(
        "SELECT form_version_id, form_no, name, revised_date, effective_from, source"
        " FROM form_versions ORDER BY form_no").fetchall())
    con.close()
    return rows


# 실제 서식 엑셀(원본) 경로 — B안: 원본에 데이터만 채워 제공
# fund-erp/templates/ 를 먼저 찾고(자립 배포), 없으면 상위 폴더(02_프로그램)로 폴백.
def _find_template(fname):
    cand = os.path.join(BASE, "templates", fname)
    if os.path.exists(cand):
        return cand
    return os.path.join(os.path.dirname(BASE), fname)


SETUP_XLSX = _find_template("2024년 근로복지기금 설립인가신청서 양식_240906.xlsx")


@app.get("/api/funds/{fund_id}/setup-excel")
def setup_excel(fund_id: str):
    """B안: 실제 설립 서식 엑셀(6종 시트, VLOOKUP)의 기금법인정보 3행을
    이 기금 데이터로 채워 반환 → 사용자가 열면 6종 서류가 실제 서식대로 완성됨."""
    if not os.path.exists(SETUP_XLSX):
        raise HTTPException(404, "원본 서식 엑셀을 찾을 수 없습니다 (02_프로그램 폴더)")
    con = connect()
    f = con.execute("SELECT * FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    fund = dict(f)
    prof = json.loads(fund.get("profile") or "{}")
    sites = rows_to_dicts(con.execute(
        "SELECT name FROM sites WHERE fund_id=? AND status='active' ORDER BY seq_label",
        (fund_id,)).fetchall())
    con.close()

    def first(txt):
        return [x.strip() for x in re.split(r"[,/·]", str(txt or "")) if x.strip()]
    wk = first(prof.get("worker_committee"))
    er = first(prof.get("emp_committee"))
    rep_site = fund.get("rep_org") or (sites[0]["name"] if sites else "")
    yd = (prof.get("years") or {})
    budget = prof.get("contribution_total")

    wb = openpyxl.load_workbook(SETUP_XLSX, data_only=False)
    ws = wb["기금법인정보"]
    # 3행 = 연번 1 (서류 시트가 VLOOKUP으로 참조하는 입력부)
    ws["B3"] = rep_site
    ws["C3"] = fund.get("name") or ""
    ws["D3"] = fund.get("chairman") or ""
    ws["E3"] = fund.get("address") or ""
    ws["F3"] = fund.get("inka_date") or ""
    ws["G3"] = prof.get("worker_rep") or (wk[0] if wk else "")
    ws["H3"] = wk[0] if wk else ""
    ws["I3"] = er[0] if er else ""
    ws["J3"] = fund.get("phone") or ""
    for cell in ("K3", "L3", "M3", "N3"):     # 생년월일 — DB에 없음 → 공란(샘플 날짜 제거)
        ws[cell] = ""
    # 출연금액·작성일자
    if budget:
        try:
            ws["F13"] = f"{int(budget):,}원"
        except (TypeError, ValueError):
            ws["F13"] = str(budget)
    if prof.get("estab_date"):
        ws["B23"] = prof["estab_date"]
    if prof.get("meeting_date"):
        ws["D23"] = prof["meeting_date"]
    # 회의록 샘플 안건(다른 기금 내용) 제거 → 사용자가 엑셀에서 직접 입력
    for r in range(33, 39):
        for col in "BCDEFGHIJKLMNOP":
            c = ws[f"{col}{r}"]
            if isinstance(c.value, str) and not c.value.startswith("="):
                c.value = ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    nm = (fund.get("short_name") or fund.get("name") or fund_id).replace(" ", "")
    fname = f"{nm}_설립서식(데이터채움).xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


# 지원신청서 원본 서식 경로 — B안: 원본에 데이터만 채워 제공
SUBSIDY_XLSX = _find_template("2025년 공동근로복지기금지원신청서 양식(폐업 사업장 미포함).xlsx")


@app.get("/api/funds/{fund_id}/subsidy-excel")
def subsidy_excel(fund_id: str):
    """B안: 실제 지원신청서 엑셀(신청서 시트+VLOOKUP)의 기금법인정보/참여사업장정보를
    이 기금 데이터로 채우고 연번 1로 고정해 반환 → 신청서 서류가 실제 서식대로 완성됨.
    서술형 칸(사업목적·추진경위·현황특성)은 사용자가 엑셀에서 직접 작성."""
    if not os.path.exists(SUBSIDY_XLSX):
        raise HTTPException(404, "지원신청서 원본 서식 엑셀을 찾을 수 없습니다 (02_프로그램 폴더)")
    con = connect()
    f = con.execute("SELECT * FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    fund = dict(f)
    if (fund.get("fund_type") or "") != "공동":
        con.close()
        raise HTTPException(400, "지원신청서는 공동근로복지기금 전용입니다")
    prof = json.loads(fund.get("profile") or "{}")
    sites = rows_to_dicts(con.execute(
        "SELECT * FROM sites WHERE fund_id=? AND status='active' ORDER BY seq_label",
        (fund_id,)).fetchall())
    con.close()

    yd = (prof.get("years") or {})
    # 최신 연도의 지원금·근로자수 정보(있으면)
    ylatest = {}
    for _y in sorted(yd.keys(), reverse=True):
        if isinstance(yd[_y], dict):
            ylatest = yd[_y]
            break
    sub = ylatest.get("subsidy") or {}

    def num(v):
        try:
            return int(str(v).replace(",", "").replace("원", "").strip())
        except (TypeError, ValueError, AttributeError):
            return None

    worker_total = sub.get("worker_count") or sum(
        (num(s.get("company_size")) or 0) for s in sites) or None
    contribution = num(prof.get("contribution_total")) or num(sub.get("contribution"))
    apply_amt = num(sub.get("apply_amount")) or contribution

    wb = openpyxl.load_workbook(SUBSIDY_XLSX, data_only=False)
    info = wb["기금법인정보"]
    # 2행 = 연번 1 (신청서 시트가 VLOOKUP AM3=1 로 참조하는 입력부)
    info["A2"] = 1
    info["B2"] = fund.get("name") or ""
    info["C2"] = fund.get("chairman") or ""
    info["D2"] = fund.get("address") or ""
    info["E2"] = fund.get("phone") or ""
    info["F2"] = fund.get("inka_no") or ""
    info["G2"] = fund.get("corp_reg_no") or ""
    info["H2"] = fund.get("tax_id_no") or ""
    info["I2"] = fund.get("reg_date") or fund.get("inka_date") or ""
    info["J2"] = len(sites) or ""
    info["K2"] = worker_total or ""
    info["L2"] = contribution or ""
    info["N2"] = apply_amt or ""
    info["S2"] = fund.get("rep_org") or ""
    info["T2"] = sub.get("author") or fund.get("manager") or ""
    # 서술형 칸은 사용자 작성 → 비움
    for col in ("P", "Q", "R"):
        info[f"{col}2"] = ""
    # 다른 기금(2~8행) 샘플 데이터 제거
    for r in range(3, 9):
        for col in "BCDEFGHIJKLMNOPQRST":
            c = info[f"{col}{r}"]
            if not (isinstance(c.value, str) and str(c.value).startswith("=")):
                c.value = None

    # 참여사업장정보 채움 (헤더 2행, 데이터 3행~), 연번 1 → "1-N"
    parts = wb["참여사업장정보"]
    start = 3
    for i, s in enumerate(sites):
        r = start + i
        # 담당자 연락처(contacts JSON) 방어적 파싱
        cname = cphone = cmail = ""
        try:
            cts = json.loads(s.get("contacts") or "[]")
            if isinstance(cts, list) and cts:
                c0 = cts[0]
                cname = c0.get("name", "")
                cphone = c0.get("phone", "") or c0.get("mobile", "")
                cmail = c0.get("email", "")
        except (ValueError, AttributeError, TypeError):
            pass
        parts.cell(r, 1, f"1-{i+1}")            # A 기금법인참여사
        parts.cell(r, 2, i + 1)                 # B 연번
        parts.cell(r, 3, s.get("name") or "")   # C 상호
        parts.cell(r, 4, s.get("ceo") or "")    # D 대표자
        parts.cell(r, 5, s.get("address") or "")   # E 소재지
        parts.cell(r, 6, num(s.get("company_size")) or "")  # F 상시근로자수
        parts.cell(r, 7, s.get("biz_type") or "")   # G 업종
        parts.cell(r, 9, s.get("biz_no") or "")     # I 사업자등록번호
        parts.cell(r, 10, num(s.get("company_size")) or "")  # J 인원
        parts.cell(r, 11, cname)                # K 담당자
        parts.cell(r, 14, cphone)               # N 휴대폰번호
        parts.cell(r, 15, cmail)                # O 이메일
    # 남은 샘플 행 제거 (다른 기금 사업장)
    for r in range(start + len(sites), parts.max_row + 1):
        for col in range(1, 21):
            c = parts.cell(r, col)
            if not (isinstance(c.value, str) and str(c.value).startswith("=")):
                c.value = None

    # 신청서 시트: 연번 1 고정 + 참여사업장 키셀(I열 "N-n")을 "1-n"으로 재작성
    doc = wb["신청서"]
    doc["AM3"] = 1
    import re as _re
    keycells = sorted(
        [c for row in doc.iter_rows(min_col=9, max_col=9) for c in row
         if isinstance(c.value, str) and _re.match(r"^\d+-\d+$", c.value.strip())],
        key=lambda c: c.row)
    for idx, c in enumerate(keycells):
        c.value = f"1-{idx+1}" if idx < len(sites) else ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    nm = (fund.get("short_name") or fund.get("name") or fund_id).replace(" ", "")
    fname = f"{nm}_지원신청서(데이터채움).xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


# 서식 자료실 카탈로그 — 단계별로 모든 서식 정리 (docgen kind + 공식 원본)
FORM_CATALOG = [
    {"cat": "① 노동부 인가 (설립)", "items": [
        {"kind": "inka", "name": "설립인가신청서", "official": "별지 제7호서식"},
        {"kind": "charter", "name": "정관"},
        {"kind": "agreement", "name": "설립합의서", "gong": True},
        {"kind": "contrib", "name": "기금출연확인서"},
        {"kind": "minutes", "name": "설립준비위원회 회의록"},
        {"kind": "bizplan", "name": "사업계획서"},
    ]},
    {"cat": "② 법원 등기", "items": [
        {"kind": "reg_apply", "name": "특수법인 설립등기신청서"},
        {"kind": "reg_accept", "name": "취임승낙서"},
        {"kind": "reg_roster", "name": "협의회 명부"},
        {"kind": "reg_seal", "name": "법인인감·개인 신고서"},
        {"kind": "reg_sealpaper", "name": "인감대지"},
        {"kind": "reg_sealcard", "name": "인감카드 (재)발급신청서"},
        {"kind": "reg_license", "name": "등록면허세 신고서"},
        {"kind": "reg_proxy", "name": "위임장(등기)"},
    ]},
    {"cat": "③ 정부지원금 (공단·지자체)", "items": [
        {"kind": "subsidy", "name": "공동근로복지기금 지원신청서", "official": "별지 제1호의2서식", "gong": True},
    ]},
    {"cat": "④ 운영·결산·보고", "items": [
        {"kind": "settlement", "name": "결산서(재무제표)"},
        {"kind": "form15", "name": "운영상황보고서", "official": "별지 제15호서식"},
    ]},
    {"cat": "⑤ 변경·해산 (공식 원본)", "items": [
        {"kind": "", "name": "기본재산 총액 변경 내용 보고서", "official": "별지 제10호서식"},
        {"kind": "", "name": "정관변경 인가신청서", "official": "별지 제11호서식"},
        {"kind": "", "name": "해산통지서", "official": "별지 제14호서식"},
    ]},
]


@app.get("/api/forms/catalog")
def forms_catalog():
    """서식 자료실 카탈로그 — 단계별 서식 목록 + 공식 원본 매칭"""
    con = connect()
    fv = {r["form_no"]: r["form_version_id"] for r in rows_to_dicts(
        con.execute("SELECT form_no, form_version_id FROM form_versions").fetchall())}
    con.close()
    out = []
    for grp in FORM_CATALOG:
        items = []
        for it in grp["items"]:
            fid = None
            if it.get("official"):
                # "별지 제7호서식" → form_versions에서 매칭 (호 번호 기준)
                m = re.search(r"제(\d+)호", it["official"])
                if m:
                    for fno, vid in fv.items():
                        if f"제{int(m.group(1))}호" in fno or f"제 {int(m.group(1))}호" in fno:
                            fid = vid
                            break
            items.append({**it, "form_version_id": fid,
                          "editable": bool(it.get("kind"))})
        out.append({"cat": grp["cat"], "items": items})
    return out


class DocEditBody(BaseModel):
    html: str


@app.put("/api/funds/{fund_id}/docedit/{kind}")
def save_docedit(fund_id: str, kind: str, body: DocEditBody, year: int = datetime.now().year):
    con = connect()
    con.execute("INSERT INTO doc_edits(fund_id,year,kind,html) VALUES(?,?,?,?)"
                " ON CONFLICT(fund_id,year,kind) DO UPDATE SET html=excluded.html,"
                " updated_at=datetime('now','localtime')", (fund_id, year, kind, body.html))
    audit(con, "doc_edits", f"{fund_id}/{year}/{kind}", "update", after="서식 편집 저장")
    con.commit()
    con.close()
    return {"ok": True}


@app.delete("/api/funds/{fund_id}/docedit/{kind}")
def reset_docedit(fund_id: str, kind: str, year: int = datetime.now().year):
    con = connect()
    con.execute("DELETE FROM doc_edits WHERE fund_id=? AND year=? AND kind=?", (fund_id, year, kind))
    con.commit()
    con.close()
    return {"ok": True}


@app.get("/api/formfile/{form_version_id}")
def form_file(form_version_id: str, ext: str = "pdf"):
    if ext not in ("hwp", "pdf"):
        raise HTTPException(400, "ext는 hwp|pdf")
    con = connect()
    row = con.execute("SELECT form_no FROM form_versions WHERE form_version_id=?",
                      (form_version_id,)).fetchone()
    con.close()
    if not row:
        raise HTTPException(404, "서식 없음")
    m = re.search(r"(\d+)", row["form_no"])
    if not m:
        raise HTTPException(404, "서식번호 해석 불가")
    key = f"별지제{int(m.group(1)):02d}호"
    if not os.path.isdir(FORMS_DIR):
        raise HTTPException(404, f"서식 폴더 없음: 00_표준서식/01_공식법정서식_현행")
    for fn in os.listdir(FORMS_DIR):
        if key in fn and fn.lower().endswith("." + ext):
            return FileResponse(os.path.join(FORMS_DIR, fn), filename=fn)
    raise HTTPException(404, f"{key} {ext.upper()} 파일 없음")


# ═══════════ 업무 보드 (B안 홈) — 업무별 파이프라인에 기금 배치 ═══════════
BOARD_JOBS = {
    "report":  {"title": "운영상황보고", "tab": "closing",
                "stages": ["자료 수령 전", "분개 승인", "결산 마감 대기", "보고서 작성", "제출 완료"]},
    "renewal": {"title": "연간 갱신(사업장)", "tab": "sites",
                "stages": ["미시작", "갱신 중", "확정 대기", "잠금 완료"]},
    "subsidy": {"title": "공단·지자체 지원금", "tab": "subsidy",
                "stages": ["미등록", "신청", "결정·교부", "정산"]},
}


def _fund_job_stage(con, f, job, year):
    """단일 기금의 업무별 현재 단계 판정 → (stage, detail, warn) 또는 None(대상 아님).
    board(전 기금 비교)와 cycle(단일 기금 흐름)이 공유."""
    fid = f["fund_id"]
    detail, warn = "", False
    if job == "report":
        agg = con.execute(
            "SELECT COUNT(*) t,"
            " COALESCE(SUM(h.status='unclassified' AND t.flag!='transfer'),0) unc,"
            " COALESCE(SUM(h.status='proposed' AND t.flag!='transfer'),0) prop"
            " FROM journal_headers h JOIN bank_transactions t ON t.tx_id=h.tx_id"
            " WHERE h.fund_id=? AND substr(h.je_date,1,4)=?", (fid, str(year))).fetchone()
        locked = con.execute("SELECT 1 FROM closing_periods WHERE fund_id=? AND year=?"
                             " AND status='locked'", (fid, year)).fetchone()
        rpt = con.execute("SELECT 1 FROM tasks WHERE fund_id=? AND year=? AND code='RPT-03'"
                          " AND status='done'", (fid, year)).fetchone()
        if rpt:
            stage = "제출 완료"
        elif locked:
            stage = "보고서 작성"
        elif agg["t"] > 0 and agg["unc"] == 0 and agg["prop"] == 0:
            stage = "결산 마감 대기"
        elif agg["t"] > 0:
            stage = "분개 승인"
            detail = f"미분류 {agg['unc']}·미승인 {agg['prop']}"
            warn = agg["unc"] > 0
        else:
            stage = "자료 수령 전"
    elif job == "renewal":
        s = con.execute("SELECT COUNT(*) FROM sites WHERE fund_id=? AND status='active'",
                        (fid,)).fetchone()[0]
        if s == 0:
            return None                       # 참여사업장 없는 기금은 갱신 대상 아님
        agg = con.execute(
            "SELECT COUNT(*) snap, COALESCE(SUM(h.locked),0) lk,"
            " COALESCE(SUM(CASE WHEN st.status='active' AND (h.worker_rep IS NULL OR"
            " h.worker_rep='' OR h.employees IS NULL) THEN 1 ELSE 0 END),0) miss"
            " FROM site_histories h JOIN sites st ON st.site_id=h.site_id"
            " WHERE st.fund_id=? AND h.year=?", (fid, year)).fetchone()
        detail = f"{min(agg['snap'], s)}/{s}" + (f" · 누락 {agg['miss']}" if agg["miss"] else "")
        warn = agg["miss"] > 0
        if agg["snap"] == 0:
            stage = "미시작"
        elif agg["lk"] >= s and agg["miss"] == 0:
            stage = "잠금 완료"
        elif agg["snap"] >= s and agg["miss"] == 0:
            stage = "확정 대기"
        else:
            stage = "갱신 중"
    else:  # subsidy
        if f["fund_type"] != "공동":
            return None
        try:
            sup = (json.loads(f["profile"] or "{}").get("years") or {}).get(str(year), {}).get("subsidy") or {}
        except Exception:
            sup = {}
        def _amt(k):
            return re.sub(r"[^0-9]", "", str(sup.get(k) or ""))
        if _amt("spent_amount"):
            stage = "정산"
            detail = f"집행 {int(_amt('spent_amount')):,}"
        elif _amt("decided_amount") or sup.get("paid_date"):
            stage = "결정·교부"
            detail = f"결정 {int(_amt('decided_amount') or 0):,}"
        elif _amt("request_amount"):
            stage = "신청"
            detail = f"신청 {int(_amt('request_amount')):,}"
        else:
            stage = "미등록"
    return stage, detail, warn


@app.get("/api/board")
def board(job: str = "report", year: int = datetime.now().year):
    if job not in BOARD_JOBS:
        raise HTTPException(400, "job은 report|renewal|subsidy")
    meta = BOARD_JOBS[job]
    con = connect()
    funds = rows_to_dicts(con.execute(
        "SELECT fund_id, name, short_name, fund_type, region, manager, status, mgmt_type, profile"
        " FROM funds WHERE mgmt_type NOT IN ('기록보관','계약종료') OR mgmt_type IS NULL OR mgmt_type=''"
        " ORDER BY region DESC, CAST(replace(replace(short_name,'충남 ',''),'호','') AS INTEGER)"
    ).fetchall())
    stages = {s: [] for s in meta["stages"]}
    for f in funds:
        res = _fund_job_stage(con, f, job, year)
        if res is None:
            continue
        stage, detail, warn = res
        stages[stage].append({
            "fund_id": f["fund_id"], "short_name": f["short_name"] or f["name"],
            "manager": f["manager"] or "", "warn": warn, "detail": detail})
    con.close()
    return {"job": job, "title": meta["title"], "tab": meta["tab"], "year": year,
            "stages": [{"key": s, "funds": stages[s]} for s in meta["stages"]],
            "total": sum(len(v) for v in stages.values()),
            "done": len(stages[meta["stages"][-1]])}


# 연간 사이클 5단계 — 상세 화면 상단 스테퍼 (계획→집행→지원금→결산→보고)
CYCLE_STEPS = [
    {"key": "plan",     "label": "계획·예산",   "tab": "annual"},
    {"key": "execute",  "label": "출연·집행",   "tab": "closing"},
    {"key": "subsidy",  "label": "지원금",      "tab": "subsidy", "gong": True},
    {"key": "closing",  "label": "결산·총회",   "tab": "closing"},
    {"key": "report",   "label": "보고·청구",   "tab": "closing"},
]


@app.get("/api/funds/{fund_id}/cycle")
def fund_cycle(fund_id: str, year: int = datetime.now().year):
    """단일 기금의 연간 사이클 진행 상태 — 상세 상단 스테퍼용.
    각 단계 done/current/todo 판정 + report/renewal/subsidy 원시 단계."""
    con = connect()
    f = con.execute(
        "SELECT fund_id, name, short_name, fund_type, region, manager, status, mgmt_type, profile"
        " FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    f = dict(f)
    is_gong = f["fund_type"] == "공동"
    rep = _fund_job_stage(con, f, "report", year) or ("자료 수령 전", "", False)
    ren = _fund_job_stage(con, f, "renewal", year)
    sub = _fund_job_stage(con, f, "subsidy", year) if is_gong else None
    # 계획·예산: profile.years[y].budget 존재 여부
    try:
        yprof = (json.loads(f["profile"] or "{}").get("years") or {}).get(str(year), {})
    except Exception:
        yprof = {}
    has_plan = bool(yprof.get("budget") or yprof.get("bizplan"))
    rep_stage = rep[0]
    # 각 사이클 단계 상태 산정
    def st(done, current):
        return "done" if done else ("current" if current else "todo")
    plan_done = has_plan
    exec_done = rep_stage in ("결산 마감 대기", "보고서 작성", "제출 완료")
    sub_stage = sub[0] if sub else None
    sub_done = sub_stage in ("정산",) if sub else True   # 공동 아니면 스킵(완료 취급)
    close_done = rep_stage in ("보고서 작성", "제출 완료")
    report_done = rep_stage == "제출 완료"
    steps = []
    prev_done = True
    for s in CYCLE_STEPS:
        if s.get("gong") and not is_gong:
            continue
        k = s["key"]
        done = {"plan": plan_done, "execute": exec_done, "subsidy": sub_done,
                "closing": close_done, "report": report_done}[k]
        detail = ""
        if k == "subsidy" and sub:
            detail = f"{sub_stage}" + (f" · {sub[1]}" if sub[1] else "")
        elif k in ("execute", "closing", "report"):
            detail = rep_stage
        current = (not done) and prev_done
        steps.append({"key": k, "label": s["label"], "tab": s["tab"],
                      "state": st(done, current), "detail": detail})
        prev_done = prev_done and done
    con.close()
    return {"year": year, "steps": steps,
            "raw": {"report": rep_stage, "renewal": (ren[0] if ren else None),
                    "subsidy": sub_stage}}


# ═══════════ 변경 이벤트 (수시) — 사업장탈퇴·정관변경·임원변경·주소이전·해산 ═══════════
EVENT_KINDS = {
    "site_leave": {"label": "사업장 탈퇴", "gong": True, "steps": [
        "탈퇴 사유 확인·정리", "참여사업장 탭에서 탈퇴 처리", "기본재산 총액 변경 여부 검토",
        "필요 시 기본재산 변경보고(별지 제10호)", "탈퇴 관련 서류 스캔 보관"]},
    "charter": {"label": "정관 변경", "steps": [
        "변경안 작성", "협의회(총회) 의결·회의록", "정관변경 인가신청(별지 제11호)",
        "인가증 수령", "변경등기 신청", "변경 정관·등기부 스캔 보관", "운영 서류 탭 정관 버전 갱신"]},
    "officer": {"label": "임원·이사 변경", "steps": [
        "변경 대상·사유 확인", "선임·해임 의결(회의록)", "취임승낙서 등 등기서류 준비",
        "임원 변경등기 신청", "등기 완료 확인", "임원 명부 갱신(운영 서류 탭)", "변경 서류 스캔 보관"]},
    "address": {"label": "주소 이전", "steps": [
        "이전 결정·의결", "소재지 변경등기 신청", "기금 기본정보 소재지 수정(일괄수정)",
        "관할 세무서·노동청 변경 확인", "변경 서류 스캔 보관"]},
    "dissolve": {"label": "해산·청산", "steps": [
        "해산 사유·의결", "해산 인가·신고", "해산등기", "청산 절차·잔여재산 처리",
        "해산통지서(별지 제14호)", "청산종결등기", "기록 보관 전환"]},
    "etc": {"label": "기타 변경", "steps": ["내용 확인", "처리", "관련 서류 스캔 보관"]},
}


class EventBody(BaseModel):
    kind: str = "etc"
    title: str = ""
    started_date: str = ""
    note: str = ""


class EventPatch(BaseModel):
    step_index: int | None = None   # 체크리스트 토글
    step_done: bool | None = None
    status: str | None = None       # 진행 | 완료
    note: str | None = None
    done_date: str | None = None


@app.get("/api/funds/{fund_id}/events")
def fund_events(fund_id: str):
    con = connect()
    rows = rows_to_dicts(con.execute(
        "SELECT * FROM fund_events WHERE fund_id=? ORDER BY status='완료',"
        " COALESCE(started_date,''), created_at DESC", (fund_id,)).fetchall())
    con.close()
    for r in rows:
        try:
            r["steps"] = json.loads(r.get("steps") or "[]")
        except Exception:
            r["steps"] = []
        r["kind_label"] = EVENT_KINDS.get(r["kind"], {}).get("label", r["kind"])
        n = len(r["steps"])
        d = sum(1 for s in r["steps"] if s.get("done"))
        r["progress"] = {"done": d, "total": n}
    return {"events": rows, "kinds": [{"key": k, "label": v["label"], "gong": v.get("gong", False)}
                                      for k, v in EVENT_KINDS.items()]}


@app.post("/api/funds/{fund_id}/events")
def create_event(fund_id: str, body: EventBody):
    meta = EVENT_KINDS.get(body.kind)
    if not meta:
        raise HTTPException(400, "알 수 없는 변경 유형")
    con = connect()
    if not con.execute("SELECT 1 FROM funds WHERE fund_id=?", (fund_id,)).fetchone():
        con.close()
        raise HTTPException(404, "기금 없음")
    eid = next_id(con, "fund_events", "event_id", "EVT")
    steps = json.dumps([{"label": s, "done": False} for s in meta["steps"]], ensure_ascii=False)
    now = datetime.now().isoformat(timespec="seconds")
    con.execute(
        "INSERT INTO fund_events(event_id, fund_id, kind, title, status, started_date,"
        " steps, note, created_at) VALUES(?,?,?,?,'진행',?,?,?,?)",
        (eid, fund_id, body.kind, body.title or meta["label"],
         body.started_date or now[:10], steps, body.note, now))
    con.commit()
    con.close()
    return {"ok": True, "event_id": eid}


@app.patch("/api/funds/{fund_id}/events/{event_id}")
def patch_event(fund_id: str, event_id: str, body: EventPatch):
    con = connect()
    row = con.execute("SELECT * FROM fund_events WHERE event_id=? AND fund_id=?",
                      (event_id, fund_id)).fetchone()
    if not row:
        con.close()
        raise HTTPException(404, "이벤트 없음")
    row = dict(row)
    steps = json.loads(row.get("steps") or "[]")
    if body.step_index is not None and 0 <= body.step_index < len(steps):
        steps[body.step_index]["done"] = bool(body.step_done)
    status = body.status if body.status is not None else row["status"]
    done_date = row["done_date"]
    if status == "완료" and not done_date:
        done_date = body.done_date or datetime.now().isoformat(timespec="seconds")[:10]
    if status == "진행":
        done_date = None
    note = body.note if body.note is not None else row["note"]
    con.execute(
        "UPDATE fund_events SET steps=?, status=?, done_date=?, note=? WHERE event_id=?",
        (json.dumps(steps, ensure_ascii=False), status, done_date, note, event_id))
    con.commit()
    con.close()
    return {"ok": True}


@app.delete("/api/funds/{fund_id}/events/{event_id}")
def delete_event(fund_id: str, event_id: str):
    con = connect()
    con.execute("DELETE FROM fund_events WHERE event_id=? AND fund_id=?", (event_id, fund_id))
    con.commit()
    con.close()
    return {"ok": True}


# ═══════════ 푸른이알피(pu-erp) 연동 — JSON 백업 가져오기 ═══════════
def _norm_name(s):
    return re.sub(r"\s|\(주\)|㈜|주식회사", "", str(s or ""))


def _find_puerp_companies(data):
    """pu-erp 백업 JSON에서 업체 배열과 사용자 배열을 탐색 (형태 유연 대응)"""
    companies, users = None, None
    def looks_company(arr):
        return (isinstance(arr, list) and arr and isinstance(arr[0], dict)
                and "name" in arr[0] and ("typeCode" in arr[0] or "bizNo" in arr[0]))
    def looks_users(arr):
        return (isinstance(arr, list) and arr and isinstance(arr[0], dict)
                and "name" in arr[0] and ("sid" in arr[0] or "id" in arr[0])
                and "typeCode" not in arr[0] and "bizNo" not in arr[0])
    def walk(obj):
        nonlocal companies, users
        if isinstance(obj, list):
            if companies is None and looks_company(obj):
                companies = obj
            elif users is None and looks_users(obj):
                users = obj
            return
        if isinstance(obj, dict):
            for k, v in obj.items():
                if companies is None and "compan" in k.lower() and looks_company(v):
                    companies = v
                elif users is None and "user" in k.lower() and isinstance(v, list):
                    users = v
                else:
                    walk(v)
    walk(data)
    return companies or [], users or []


@app.post("/api/import/puerp")
async def import_puerp(file: UploadFile = File(...), create_new: bool = True):
    """푸른이알피 JSON 백업 → 기금(typeCode='기금') 매칭·동기화
    - 기존 기금과 이름/대표사업장으로 매칭 → 담당자·계약기간·연락처 갱신
    - 미매칭 + 이름에 '기금' 포함 → 자문 기금으로 신규 등록(create_new)
    """
    raw = await file.read()
    try:
        data = json.loads(raw.decode("utf-8-sig"))
    except Exception:
        raise HTTPException(400, "JSON 파싱 실패 — 푸른이알피 백업(JSON) 파일인지 확인")
    companies, users = _find_puerp_companies(data)
    if not companies:
        raise HTTPException(400, "업체 목록을 찾지 못함 (typeCode/name 배열 없음)")
    uid2name, name2uid = {}, {}
    for u in users:
        for k in ("sid", "id"):
            if u.get(k):
                uid2name[str(u[k])] = u.get("name", "")
                if u.get("name"):
                    name2uid[u["name"]] = str(u[k])
    def resolve_staff(v):
        """sid면 (이름, sid), 이름이면 (이름, 매칭 sid) 반환 — 푸른이알피 연결키 보존."""
        s = str(v or "").strip()
        if s in uid2name:
            return uid2name[s], s
        return s, name2uid.get(s, "")

    con = connect()
    funds = rows_to_dicts(con.execute(
        "SELECT fund_id, name, short_name, rep_org, mgmt_type FROM funds").fetchall())
    by_name = {_norm_name(f["name"]): f for f in funds}
    by_short = {_norm_name(f["short_name"]): f for f in funds if f["short_name"]}
    by_rep = {_norm_name(f["rep_org"]): f for f in funds if f["rep_org"]}

    matched, created, updated, skipped = [], [], [], []
    for co in companies:
        if co.get("typeCode") != "기금" and "기금" not in str(co.get("name", "")):
            continue                      # 기금 관련 업체만
        nm = str(co.get("name", "")).strip()
        key = _norm_name(nm)
        f = by_name.get(key) or by_short.get(key) or by_rep.get(key)
        if not f:  # 부분 일치
            f = next((x for x in funds if key and (_norm_name(x["name"]).find(key) >= 0
                                                   or key.find(_norm_name(x["short_name"] or "§")) >= 0)), None)
        main, main_uid = resolve_staff(co.get("managerMain"))
        sub_pairs = [resolve_staff(s) for s in (co.get("managerSubs") or [])]
        subs = [nm for nm, _ in sub_pairs]
        staff_uids = {}
        if main:
            staff_uids[main] = main_uid
        for nm2, uid2 in sub_pairs:
            if nm2:
                staff_uids[nm2] = uid2
        contact = (co.get("contacts") or [{}])
        c0 = contact[0] if contact else {}
        rep_contact = " / ".join(x for x in [
            c0.get("name", "") or co.get("primaryContactName", ""),
            c0.get("phone", "") or co.get("phone", ""),
            c0.get("email", "") or co.get("email", "")] if x)
        term = ""
        if co.get("contractStartDate") or co.get("contractEndDate"):
            term = f"계약 {co.get('contractStartDate','')}~{co.get('contractEndDate','')}"
        if f:
            fid = f["fund_id"]
            if main or subs:
                _set_fund_staff(con, fid, main, subs[0] if subs else "", staff_uids)
            if rep_contact:
                con.execute("UPDATE funds SET rep_contact=? WHERE fund_id=?", (rep_contact, fid))
            if term:
                con.execute("UPDATE funds SET note=CASE WHEN instr(note,?)>0 THEN note"
                            " ELSE trim(note||' '||?) END WHERE fund_id=?", (term, term, fid))
            audit(con, "funds", fid, "update", "puerp연동", "", nm)
            (matched if f["mgmt_type"] != "기록보관" else skipped).append(nm)
            updated.append(nm)
        elif create_new and "기금" in nm:
            fid = next_id(con, "funds", "fund_id", "FUND")
            ftype = "사내" if "사내" in nm else "공동"
            con.execute("INSERT INTO funds(fund_id,name,short_name,fund_type,region,status,"
                        "mgmt_type,advisory,rep_contact,note) VALUES(?,?,?,?,?,?,?,?,?,?)",
                        (fid, nm, nm[:12], ftype, "", "운영", "자문", "자문", rep_contact,
                         f"푸른이알피 연동 등록 {term}".strip()))
            if main or subs:
                _set_fund_staff(con, fid, main, subs[0] if subs else "", staff_uids)
            pnp = ensure_agency_app(con, "푸른노무법인", "수탁법인")
            con.execute("INSERT OR IGNORE INTO fund_agencies(fund_id,agency_id,role)"
                        " VALUES(?,?,'수탁기관')", (fid, pnp))
            audit(con, "funds", fid, "create", after=f"{nm} (푸른이알피 연동)")
            created.append(nm)
        else:
            skipped.append(nm)
    con.commit()
    con.close()
    return {"ok": True, "scanned": len(companies),
            "matched": matched, "created": created, "skipped": skipped}


# ═══════════ 임원·협의회 명부·임기 (M4) ═══════════
@app.get("/api/funds/{fund_id}/roles")
def list_roles(fund_id: str):
    con = connect()
    rows = rows_to_dicts(con.execute(
        "SELECT * FROM fund_roles WHERE fund_id=? AND status='active'"
        " ORDER BY CASE role WHEN '이사장' THEN 1 WHEN '이사' THEN 2 WHEN '감사' THEN 3"
        " WHEN '근로자위원' THEN 4 WHEN '사용자위원' THEN 5 ELSE 6 END, role_id", (fund_id,)).fetchall())
    con.close()
    today = datetime.now().strftime("%Y-%m-%d")
    from datetime import date as _date
    for r in rows:
        r["expiring"] = False
        r["expired"] = False
        if r["term_end"]:
            try:
                dd = (_date.fromisoformat(r["term_end"]) - _date.fromisoformat(today)).days
                r["days_left"] = dd
                r["expired"] = dd < 0
                r["expiring"] = 0 <= dd <= 90
            except ValueError:
                pass
    return rows


class RoleBody(BaseModel):
    person_name: str
    role: str = "이사"
    term_start: str = ""
    term_end: str = ""
    note: str = ""


@app.post("/api/funds/{fund_id}/roles")
def add_role(fund_id: str, body: RoleBody):
    con = connect()
    con.execute("INSERT INTO fund_roles(fund_id,person_name,role,term_start,term_end,note,status)"
                " VALUES(?,?,?,?,?,?, 'active')",
                (fund_id, body.person_name, body.role, body.term_start, body.term_end, body.note))
    audit(con, "fund_roles", fund_id, "create", after=f"{body.role} {body.person_name}")
    con.commit()
    con.close()
    return {"ok": True}


@app.delete("/api/roles/{role_id}")
def del_role(role_id: int):
    con = connect()
    con.execute("UPDATE fund_roles SET status='퇴임' WHERE role_id=?", (role_id,))
    audit(con, "fund_roles", str(role_id), "update", "status", "active", "퇴임")
    con.commit()
    con.close()
    return {"ok": True}


# ═══════════ 운영 서류: 회의록·정관·운영규정 (M15) ═══════════
@app.get("/api/funds/{fund_id}/meetings")
def list_meetings(fund_id: str):
    con = connect()
    out = rows_to_dicts(con.execute(
        "SELECT * FROM meetings WHERE fund_id=? ORDER BY mdate DESC, meeting_id DESC",
        (fund_id,)).fetchall())
    con.close()
    return out


class MeetingBody(BaseModel):
    mdate: str = ""
    mtype: str = "이사회"
    agenda: str = ""
    resolution: str = ""
    attendees: str = ""
    quorum_note: str = ""
    tag: str = ""


@app.post("/api/funds/{fund_id}/meetings")
def add_meeting(fund_id: str, body: MeetingBody):
    con = connect()
    mid = next_id(con, "meetings", "meeting_id", "MTG")
    con.execute("INSERT INTO meetings(meeting_id,fund_id,mdate,mtype,agenda,resolution,"
                "attendees,quorum_note,tag) VALUES(?,?,?,?,?,?,?,?,?)",
                (mid, fund_id, body.mdate, body.mtype, body.agenda, body.resolution,
                 body.attendees, body.quorum_note, body.tag))
    audit(con, "meetings", mid, "create", after=f"{body.mdate} {body.mtype}")
    con.commit()
    con.close()
    return {"ok": True, "meeting_id": mid}


@app.delete("/api/meetings/{meeting_id}")
def del_meeting(meeting_id: str):
    con = connect()
    con.execute("DELETE FROM meetings WHERE meeting_id=?", (meeting_id,))
    audit(con, "meetings", meeting_id, "update", "delete", "", "삭제")
    con.commit()
    con.close()
    return {"ok": True}


@app.get("/api/funds/{fund_id}/revisions")
def list_revisions(fund_id: str):
    con = connect()
    rows = rows_to_dicts(con.execute(
        "SELECT * FROM doc_revisions WHERE fund_id=? ORDER BY kind, rule_name, version DESC",
        (fund_id,)).fetchall())
    con.close()
    return rows


class RevisionBody(BaseModel):
    kind: str = "정관"
    rule_name: str = ""
    version: int = 1
    rev_date: str = ""
    inka_date: str = ""
    basis_meeting: str = ""
    note: str = ""


@app.post("/api/funds/{fund_id}/revisions")
def add_revision(fund_id: str, body: RevisionBody):
    con = connect()
    rid = next_id(con, "doc_revisions", "rev_id", "REV")
    # 같은 종류·규정명의 기존 현행본 해제 → 새 버전이 현행
    con.execute("UPDATE doc_revisions SET is_current=0 WHERE fund_id=? AND kind=? AND rule_name=?",
                (fund_id, body.kind, body.rule_name))
    con.execute("INSERT INTO doc_revisions(rev_id,fund_id,kind,rule_name,version,rev_date,"
                "inka_date,basis_meeting,is_current,note) VALUES(?,?,?,?,?,?,?,?,1,?)",
                (rid, fund_id, body.kind, body.rule_name, body.version, body.rev_date,
                 body.inka_date, body.basis_meeting, body.note))
    audit(con, "doc_revisions", rid, "create", after=f"{body.kind} {body.rule_name} v{body.version}")
    con.commit()
    con.close()
    return {"ok": True, "rev_id": rid}


# ═══════════ 연간 운영 일정·체크리스트 (M5, 계획서 §4) ═══════════
# (code, 시기, 제목, 상대기한: (연도기준 offset개월, 일) 또는 None, 공동전용 여부)
ANNUAL_TEMPLATE = [
    ("OPS-01", "매월 1~10일", "전월 통장거래 가져오기 → 자동분개 검토·승인", None, False),
    ("OPS-02", "매월 말", "월마감 (미분류·증빙 확인 후 잠금)", None, False),
    ("OPS-03", "분기", "분기 운영점검 (기본재산 변동·증빙 누락 점검)", None, False),
    ("OPS-04", "수시", "출연 약정·입금 대사 / 기본재산 변경 시 보고(별지 제10호)", None, False),
    ("SUP-01", "공고 시", "공단·지자체 지원금 신청·정산", None, True),
    ("RPT-01", "9~11월", "차기연도 사업계획·예산 수립, 협의회 상정", (0, "11-30"), False),
    ("CLS-01", "12월", "가결산 (연말 잔액·미지급 정리)", (0, "12-31"), False),
    ("CLS-02", "익년 1~2월", "결산 확정·재무제표 작성", (1, "02-15"), False),
    ("AUD-01", "익년 2월", "감사 실시·감사보고서 작성", (1, "02-20"), False),
    ("RPT-02", "익년 2~3월", "협의회 결산·차기 사업계획 승인", (1, "03-15"), False),
    ("RPT-03", "익년 3월", "운영상황보고서(별지 제15호) 제출 — 회계연도 종료 후 3개월 이내", (1, "03-31"), False),
    ("TAX-01", "익년 3월", "수익사업·원천세 신고 해당성 검토", (1, "03-31"), False),
]


def _seed_annual(con, fund_id, year, ftype):
    exists = con.execute("SELECT COUNT(*) FROM tasks WHERE fund_id=? AND year=?",
                         (fund_id, year)).fetchone()[0]
    if exists:
        return
    for i, (code, period, title, due, gong) in enumerate(ANNUAL_TEMPLATE):
        if gong and ftype != "공동":
            continue
        due_date = ""
        if due:
            due_date = f"{year + due[0]}-{due[1]}"
        # 시기(템플릿) 순서 보존을 위해 task_id에 2자리 순번 부여
        con.execute("INSERT INTO tasks(task_id,fund_id,year,code,period,title,due_date,status)"
                    " VALUES(?,?,?,?,?,?,?, 'todo')",
                    (f"TASK-{fund_id}-{year}-{i:02d}-{code}", fund_id, year, code, period, title, due_date))


@app.get("/api/funds/{fund_id}/annual/{year}")
def annual(fund_id: str, year: int):
    con = connect()
    f = con.execute("SELECT fund_type FROM funds WHERE fund_id=?", (fund_id,)).fetchone()
    if not f:
        con.close()
        raise HTTPException(404, "기금 없음")
    _seed_annual(con, fund_id, year, f["fund_type"])
    con.commit()
    tasks = rows_to_dicts(con.execute(
        "SELECT task_id, code, period, title, due_date, status, done_date, note"
        " FROM tasks WHERE fund_id=? AND year=? ORDER BY task_id", (fund_id, year)).fetchall())
    con.close()
    done = sum(1 for t in tasks if t["status"] == "done")
    return {"year": year, "tasks": tasks, "done": done, "total": len(tasks)}


class TaskToggle(BaseModel):
    status: str = "done"
    done_date: str = ""
    note: str = ""


@app.patch("/api/tasks/{task_id}")
def patch_task(task_id: str, body: TaskToggle):
    con = connect()
    con.execute("UPDATE tasks SET status=?, done_date=?, note=? WHERE task_id=?",
                (body.status, body.done_date, body.note, task_id))
    audit(con, "tasks", task_id, "update", "status", "", body.status)
    con.commit()
    con.close()
    return {"ok": True}


# ═══════════ 목적사업·대부 (M10) ═══════════
WELFARE_CATEGORIES = ["생활안정자금", "장학금", "재난구호금", "체육문화활동", "주택자금",
                      "모성보호", "근로자의날", "근로복지시설", "경조사비", "기타복지비", "대부사업"]
# 별지 제15호 사업실적 행 매핑
FORM15_ROW = {
    "생활안정자금": "생활안정자금", "장학금": "장학금", "재난구호금": "재난구호금",
    "체육문화활동": "체육문화활동", "주택자금": "주택자금", "모성보호": "모성보호",
    "근로자의날": "근로자의날", "근로복지시설": "근로복지시설",
    "경조사비": "그밖의복지비", "기타복지비": "그밖의복지비", "대부사업": "대부사업",
}


@app.get("/api/funds/{fund_id}/welfare")
def list_welfare(fund_id: str, year: int = datetime.now().year):
    con = connect()
    progs = rows_to_dicts(con.execute(
        "SELECT * FROM welfare_programs WHERE fund_id=? AND year=? ORDER BY kind DESC, program_id",
        (fund_id, year)).fetchall())
    for p in progs:
        agg = con.execute(
            "SELECT COUNT(*) n, COALESCE(SUM(amount),0) amt, COALESCE(SUM(beneficiaries),0) ben,"
            " COALESCE(SUM(CASE WHEN status='paid' THEN amount ELSE 0 END),0) paid"
            " FROM welfare_cases WHERE program_id=?", (p["program_id"],)).fetchone()
        p["case_count"], p["used"], p["beneficiaries"], p["paid"] = \
            agg["n"], agg["amt"], agg["ben"], agg["paid"]
    con.close()
    return {"year": year, "programs": progs, "categories": WELFARE_CATEGORIES}


class WelfareBody(BaseModel):
    name: str
    kind: str = "목적사업"
    category: str = "기타복지비"
    budget: int = 0
    year: int


@app.post("/api/funds/{fund_id}/welfare")
def create_welfare(fund_id: str, body: WelfareBody):
    con = connect()
    pid = next_id(con, "welfare_programs", "program_id", "WP")
    con.execute("INSERT INTO welfare_programs(program_id,fund_id,year,name,kind,category,budget)"
                " VALUES(?,?,?,?,?,?,?)",
                (pid, fund_id, body.year, body.name, body.kind, body.category, body.budget))
    audit(con, "welfare_programs", pid, "create", after=body.name)
    con.commit()
    con.close()
    return {"ok": True, "program_id": pid}


class WelfarePatch(BaseModel):
    field: str
    value: str | int | None = None


@app.patch("/api/welfare/{program_id}")
def patch_welfare(program_id: str, body: WelfarePatch):
    if body.field not in ("name", "kind", "category", "budget", "note", "approved_minutes"):
        raise HTTPException(400, "수정 불가 필드")
    con = connect()
    v = body.value
    if body.field == "budget" and v is not None:
        v = int(re.sub(r"[^0-9]", "", str(v)) or 0)
    con.execute(f"UPDATE welfare_programs SET {body.field}=? WHERE program_id=?", (v, program_id))
    audit(con, "welfare_programs", program_id, "update", body.field, "", v)
    con.commit()
    con.close()
    return {"ok": True}


@app.get("/api/welfare/{program_id}/cases")
def list_cases(program_id: str):
    con = connect()
    cases = rows_to_dicts(con.execute(
        "SELECT c.*, s.name site_name FROM welfare_cases c"
        " LEFT JOIN sites s ON s.site_id=c.site_id WHERE c.program_id=?"
        " ORDER BY c.applied_date", (program_id,)).fetchall())
    con.close()
    return cases


class CaseBody(BaseModel):
    site_id: str = ""
    applied_date: str = ""
    amount: int = 0
    beneficiaries: int = 1
    status: str = "paid"
    note: str = ""


@app.post("/api/welfare/{program_id}/cases")
def add_case(program_id: str, body: CaseBody):
    con = connect()
    cid = next_id(con, "welfare_cases", "case_id", "WC")
    con.execute("INSERT INTO welfare_cases(case_id,program_id,site_id,applied_date,amount,"
                "beneficiaries,status,note) VALUES(?,?,?,?,?,?,?,?)",
                (cid, program_id, body.site_id or None, body.applied_date, body.amount,
                 body.beneficiaries, body.status, body.note))
    audit(con, "welfare_cases", cid, "create", after=f"{body.amount}원 {body.beneficiaries}명")
    con.commit()
    con.close()
    return {"ok": True, "case_id": cid}


@app.delete("/api/cases/{case_id}")
def del_case(case_id: str):
    con = connect()
    con.execute("DELETE FROM welfare_cases WHERE case_id=?", (case_id,))
    audit(con, "welfare_cases", case_id, "update", "delete", "", "삭제")
    con.commit()
    con.close()
    return {"ok": True}


def welfare_form15(con, fund_id, year):
    """목적사업 실적 → 별지15호 사업실적 행별 (금액·수혜자수) 집계"""
    rows = rows_to_dicts(con.execute(
        "SELECT p.category, COALESCE(SUM(c.amount),0) amt, COALESCE(SUM(c.beneficiaries),0) ben"
        " FROM welfare_programs p LEFT JOIN welfare_cases c ON c.program_id=p.program_id"
        " WHERE p.fund_id=? AND p.year=? GROUP BY p.category", (fund_id, year)).fetchall())
    out = {}
    for r in rows:
        key = FORM15_ROW.get(r["category"], "그밖의복지비")
        e = out.setdefault(key, {"amt": 0, "ben": 0})
        e["amt"] += r["amt"]
        e["ben"] += r["ben"]
    return out


# ═══════════ 회계·결산 (3단계, M11·M12·M13) ═══════════
import accounting as _acc


@app.get("/api/funds/{fund_id}/accounts")
def list_accounts(fund_id: str):
    con = connect()
    out = rows_to_dicts(con.execute(
        "SELECT * FROM accounts WHERE fund_id=? ORDER BY account_id", (fund_id,)).fetchall())
    con.close()
    return out


class AccountBody(BaseModel):
    bank: str = ""
    account_no_masked: str = ""
    purpose: str = "보통재산"


@app.post("/api/funds/{fund_id}/accounts")
def create_account(fund_id: str, body: AccountBody):
    con = connect()
    aid = next_id(con, "accounts", "account_id", "ACC")
    con.execute("INSERT INTO accounts(account_id,fund_id,bank,account_no_masked,purpose)"
                " VALUES(?,?,?,?,?)", (aid, fund_id, body.bank, body.account_no_masked, body.purpose))
    audit(con, "accounts", aid, "create", after=f"{body.bank} {body.purpose}")
    con.commit()
    con.close()
    return {"ok": True, "account_id": aid}


@app.post("/api/accounts/{account_id}/import")
async def import_bank(account_id: str, file: UploadFile = File(...)):
    raw = await file.read()
    h = _acc.file_hash(raw)
    con = connect()
    acc = con.execute("SELECT fund_id FROM accounts WHERE account_id=?", (account_id,)).fetchone()
    if not acc:
        con.close()
        raise HTTPException(404, "계좌 없음")
    fund_id = acc["fund_id"]
    if con.execute("SELECT 1 FROM bank_imports WHERE file_hash=?", (h,)).fetchone():
        con.close()
        raise HTTPException(409, "이미 가져온 파일입니다 (동일 내용)")
    try:
        txns = _acc.parse_bank_excel(raw)
    except ValueError as e:
        con.close()
        raise HTTPException(400, str(e))
    imp_id = next_id(con, "bank_imports", "import_id", "IMP")
    con.execute("INSERT INTO bank_imports(import_id,account_id,file_name,file_hash,imported_at,row_count)"
                " VALUES(?,?,?,?,datetime('now','localtime'),?)",
                (imp_id, account_id, file.filename, h, len(txns)))
    learned = rows_to_dicts(con.execute(
        "SELECT keyword, direction, debit, credit FROM learned_rules WHERE fund_id=?"
        " ORDER BY hits DESC", (fund_id,)).fetchall())
    n = con.execute("SELECT COUNT(*) FROM bank_transactions").fetchone()[0]
    transfer = cancel = 0
    for i, t in enumerate(txns):
        tx_id = f"TX-{datetime.now().year}-{n + i + 1:08d}"
        flag = _acc.flag_of(t["desc"])
        if flag == "transfer":
            transfer += 1
        if flag == "cancel":
            cancel += 1
        con.execute(
            "INSERT INTO bank_transactions(tx_id,import_id,account_id,tx_date,description,"
            "deposit,withdrawal,balance,flag) VALUES(?,?,?,?,?,?,?,?,?)",
            (tx_id, imp_id, account_id, t["date"], t["desc"], t["deposit"], t["withdrawal"],
             t["balance"], flag))
        # 자동분개 제안 (학습 규칙 우선)
        pj = _acc.propose_journal(t, learned)
        je_id = f"JE-{datetime.now().year}-{n + i + 1:08d}"
        con.execute("INSERT INTO journal_headers(journal_id,fund_id,tx_id,je_date,status,proposed_rule)"
                    " VALUES(?,?,?,?,?,?)",
                    (je_id, fund_id, tx_id, t["date"],
                     "proposed" if pj["auto"] else "unclassified", pj["rule"]))
        con.execute("INSERT INTO journal_lines(journal_id,side,account,amount,basic_asset)"
                    " VALUES(?,?,?,?,?)", (je_id, "D", pj["debit"], pj["amount"], pj.get("basic", 0)))
        con.execute("INSERT INTO journal_lines(journal_id,side,account,amount) VALUES(?,?,?,?)",
                    (je_id, "C", pj["credit"], pj["amount"]))
    # 계좌간이체 자동 짝짓기: 같은 기금 다른 계좌, 반대방향·동일금액·3일 이내 → 양쪽 transfer 처리
    paired = _pair_transfers(con, fund_id)
    audit(con, "bank_imports", imp_id, "import", after=f"{file.filename} {len(txns)}건")
    con.commit()
    con.close()
    return {"ok": True, "rows": len(txns), "transfer": transfer + paired, "cancel": cancel,
            "paired_transfers": paired,
            "unclassified": sum(1 for t in txns if not _acc.propose_journal(t)["auto"])}


def _pair_transfers(con, fund_id):
    """같은 기금의 서로 다른 계좌 간 반대방향·동일금액·근접일자 거래를 이체 쌍으로 연결"""
    rows = rows_to_dicts(con.execute(
        "SELECT tx_id, account_id, tx_date, deposit, withdrawal FROM bank_transactions"
        " WHERE account_id IN (SELECT account_id FROM accounts WHERE fund_id=?)"
        " AND (flag='' OR flag IS NULL) AND pair_tx_id=''", (fund_id,)).fetchall())
    from datetime import date as _d
    def pd(s):
        try:
            return _d.fromisoformat(s[:10])
        except ValueError:
            return None
    dep = [r for r in rows if r["deposit"] > 0]
    wd = [r for r in rows if r["withdrawal"] > 0]
    used = set()
    n = 0
    for a in dep:
        da = pd(a["tx_date"])
        for b in wd:
            if b["tx_id"] in used or a["account_id"] == b["account_id"]:
                continue
            if a["deposit"] != b["withdrawal"]:
                continue
            db = pd(b["tx_date"])
            if da and db and abs((da - db).days) <= 3:
                for tx in (a["tx_id"], b["tx_id"]):
                    con.execute("UPDATE bank_transactions SET flag='transfer' WHERE tx_id=?", (tx,))
                    con.execute("UPDATE journal_headers SET status='proposed',"
                                " proposed_rule='계좌간이체(자동 짝)' WHERE tx_id=?", (tx,))
                con.execute("UPDATE bank_transactions SET pair_tx_id=? WHERE tx_id=?", (b["tx_id"], a["tx_id"]))
                con.execute("UPDATE bank_transactions SET pair_tx_id=? WHERE tx_id=?", (a["tx_id"], b["tx_id"]))
                used.add(b["tx_id"])
                n += 2
                break
    return n


@app.get("/api/funds/{fund_id}/journals")
def list_journals(fund_id: str, year: int = datetime.now().year):
    con = connect()
    rows = rows_to_dicts(con.execute(
        "SELECT h.journal_id, h.je_date, h.status, h.proposed_rule, t.description, t.deposit,"
        " t.withdrawal, t.flag FROM journal_headers h JOIN bank_transactions t ON t.tx_id=h.tx_id"
        " WHERE h.fund_id=? AND substr(h.je_date,1,4)=? ORDER BY h.je_date, h.journal_id",
        (fund_id, str(year))).fetchall())
    for r in rows:
        lines = rows_to_dicts(con.execute(
            "SELECT side, account, amount FROM journal_lines WHERE journal_id=? ORDER BY side DESC",
            (r["journal_id"],)).fetchall())
        r["debit"] = next((l["account"] for l in lines if l["side"] == "D"), "")
        r["credit"] = next((l["account"] for l in lines if l["side"] == "C"), "")
        r["amount"] = lines[0]["amount"] if lines else 0
    con.close()
    return {"year": year, "journals": rows,
            "accounts": list(_acc.ACCOUNT_CHART.keys()) + ["계좌대체", "?"]}


class JournalBody(BaseModel):
    debit: str
    credit: str
    status: str = "approved"


@app.patch("/api/journals/{journal_id}")
def patch_journal(journal_id: str, body: JournalBody):
    con = connect()
    row = con.execute("SELECT fund_id, tx_id FROM journal_headers WHERE journal_id=?",
                      (journal_id,)).fetchone()
    if not row:
        con.close()
        raise HTTPException(404, "분개 없음")
    con.execute("UPDATE journal_lines SET account=? WHERE journal_id=? AND side='D'",
                (body.debit, journal_id))
    con.execute("UPDATE journal_lines SET account=? WHERE journal_id=? AND side='C'",
                (body.credit, journal_id))
    con.execute("UPDATE journal_headers SET status=?, approved_by='local',"
                " approved_at=datetime('now','localtime') WHERE journal_id=?", (body.status, journal_id))
    # 승인 시 학습: 이 거래의 적요 → (차/대) 규칙 저장 (같은 거래처 다음 자동 제안)
    if body.status == "approved":
        tx = con.execute("SELECT description, deposit, withdrawal, flag FROM bank_transactions"
                         " WHERE tx_id=?", (row["tx_id"],)).fetchone()
        if tx and tx["flag"] != "transfer":
            kw = _acc.extract_keyword(tx["description"])
            direction = "입금" if tx["deposit"] > 0 else "출금"
            if kw and body.debit != "?" and body.credit != "?":
                con.execute(
                    "INSERT INTO learned_rules(fund_id,keyword,direction,debit,credit) VALUES(?,?,?,?,?)"
                    " ON CONFLICT(fund_id,keyword,direction) DO UPDATE SET debit=excluded.debit,"
                    " credit=excluded.credit, hits=hits+1, updated_at=datetime('now','localtime')",
                    (row["fund_id"], kw, direction, body.debit, body.credit))
    audit(con, "journal_headers", journal_id, "update", "status", "", body.status)
    con.commit()
    con.close()
    return {"ok": True}


@app.post("/api/funds/{fund_id}/journals/reapply")
def reapply_learned(fund_id: str, year: int = datetime.now().year):
    """학습 규칙을 미분류·미승인(비이체) 분개에 재적용 → 제안 갱신 (확정 아님)"""
    con = connect()
    learned = rows_to_dicts(con.execute(
        "SELECT keyword, direction, debit, credit FROM learned_rules WHERE fund_id=?"
        " ORDER BY hits DESC", (fund_id,)).fetchall())
    if not learned:
        con.close()
        return {"ok": True, "updated": 0, "note": "학습된 규칙이 아직 없습니다"}
    rows = rows_to_dicts(con.execute(
        "SELECT h.journal_id, t.description, t.deposit, t.withdrawal FROM journal_headers h"
        " JOIN bank_transactions t ON t.tx_id=h.tx_id"
        " WHERE h.fund_id=? AND substr(h.je_date,1,4)=? AND h.status IN ('unclassified','proposed')"
        " AND t.flag!='transfer'", (fund_id, str(year))).fetchall())
    updated = 0
    for r in rows:
        pj = _acc.propose_journal(
            {"desc": r["description"], "deposit": r["deposit"], "withdrawal": r["withdrawal"]}, learned)
        if pj["auto"] and pj["rule"].startswith("학습:"):
            con.execute("UPDATE journal_lines SET account=? WHERE journal_id=? AND side='D'",
                        (pj["debit"], r["journal_id"]))
            con.execute("UPDATE journal_lines SET account=? WHERE journal_id=? AND side='C'",
                        (pj["credit"], r["journal_id"]))
            con.execute("UPDATE journal_headers SET status='proposed', proposed_rule=? "
                        "WHERE journal_id=?", (pj["rule"], r["journal_id"]))
            updated += 1
    audit(con, "journal_headers", fund_id, "update", "reapply", "", f"{updated}건 학습 재적용")
    con.commit()
    con.close()
    return {"ok": True, "updated": updated, "rules": len(learned)}


@app.get("/api/funds/{fund_id}/closing/{year}")
def closing(fund_id: str, year: int):
    con = connect()
    js = rows_to_dicts(con.execute(
        "SELECT h.journal_id, h.status, t.flag FROM journal_headers h"
        " JOIN bank_transactions t ON t.tx_id=h.tx_id"
        " WHERE h.fund_id=? AND substr(h.je_date,1,4)=?", (fund_id, str(year))).fetchall())
    # 계좌간이체는 수입·지출이 아니므로 관문 집계에서 제외 (시산표에서도 제외)
    real = [j for j in js if j["flag"] != "transfer"]
    unclassified = sum(1 for j in real if j["status"] == "unclassified")
    unapproved = sum(1 for j in real if j["status"] == "proposed")
    approved_ids = [j["journal_id"] for j in js
                    if j["status"] == "approved" and j["flag"] != "transfer"]
    lines = []
    if approved_ids:
        q = "SELECT side, account, amount FROM journal_lines WHERE journal_id IN (%s)" % \
            ",".join("?" * len(approved_ids))
        lines = rows_to_dicts(con.execute(q, approved_ids).fetchall())
    lock = con.execute("SELECT status FROM closing_periods WHERE fund_id=? AND year=?",
                       (fund_id, year)).fetchone()
    con.close()
    tb = _acc.trial_balance([{"side": l["side"], "account": l["account"], "amount": l["amount"]}
                             for l in lines])
    tb_d = sum(v["debit"] for v in tb.values())
    tb_c = sum(v["credit"] for v in tb.values())
    fs = _acc.financial_statements(tb)
    ready = (unclassified == 0 and unapproved == 0 and tb_d == tb_c and len(real) > 0)
    return {
        "year": year, "total_journals": len(js),
        "unclassified": unclassified, "unapproved": unapproved,
        "balanced": tb_d == tb_c, "tb_debit": tb_d, "tb_credit": tb_c,
        "trial_balance": tb, "financials": fs,
        "locked": bool(lock and lock["status"] == "locked"),
        "ready_to_close": ready,
    }


@app.post("/api/funds/{fund_id}/closing/{year}/lock")
def lock_closing(fund_id: str, year: int, unlock: bool = False):
    con = connect()
    if not unlock:
        st = closing(fund_id, year)
        if not st["ready_to_close"]:
            con.close()
            raise HTTPException(400, f"마감 불가 — 미분류 {st['unclassified']}·미승인 {st['unapproved']}"
                                     f"·대차{'일치' if st['balanced'] else '불일치'}")
    con.execute(
        "INSERT INTO closing_periods(fund_id,year,status,locked_at,locked_by) "
        "VALUES(?,?,?,datetime('now','localtime'),'local') "
        "ON CONFLICT(fund_id,year) DO UPDATE SET status=excluded.status, locked_at=excluded.locked_at",
        (fund_id, year, "open" if unlock else "locked"))
    if not unlock:
        # 결산 마감 → 보수 청구 건 자동 생성 (미청구 상태, 금액은 담당자 입력)
        _ensure_billing(con, fund_id, year, "결산·운영상황보고 보수", "결산 마감 시 자동 생성")
    audit(con, "closing_periods", f"{fund_id}/{year}", "lock", "status", "",
          "open" if unlock else "locked")
    con.commit()
    con.close()
    return {"ok": True, "locked": not unlock}


@app.get("/api/audit")
def recent_audit(limit: int = 50):
    con = connect()
    out = rows_to_dicts(con.execute(
        "SELECT * FROM audit_logs ORDER BY log_id DESC LIMIT ?", (limit,)).fetchall())
    con.close()
    return out


# ── 정적 파일 (프런트) ──
app.mount("/", StaticFiles(directory=os.path.join(BASE, "static"), html=True), name="static")
