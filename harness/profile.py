# -*- coding: utf-8 -*-
"""
하네스 1차 - 목표 1 준비: 급여대장 엑셀 구조 프로파일링
- classify 결과에서 '급여대장' + 엑셀(.xlsx/.xlsm)만 대상.
- 각 파일/시트에서 헤더행 위치와 핵심 필드 존재 여부를 파악해 '구조 시그니처'로 군집화.
- 목적: 가장 단순하고 다수인 양식을 찾아 파서 착수점 결정.
- 결과: _harness_out/profile.json, _harness_out/profile_summary.txt
"""
import os, json, re, warnings
from collections import Counter, defaultdict
warnings.filterwarnings("ignore")

DATA_ROOT = os.environ.get(
    "PAYROLL_DATA_ROOT",
    r"C:\Users\fair0\OneDrive\바탕 화면\급여아웃소싱 서류들",
)
OUT_DIR = os.path.join(DATA_ROOT, "_harness_out")

# 핵심 필드 동의어 (canonical → 검색어들)
FIELDS = {
    "성명":     ["성명", "이름", "성 명", "직원명", "근로자"],
    "기본급":   ["기본급", "기본 급"],
    "과세총액": ["과세총액", "과세대상", "과세소득", "과세금액", "과세계", "과세"],
    "소득세":   ["소득세", "갑근세"],
    "지방세":   ["지방소득세", "지방세", "주민세"],
    "국민연금": ["국민연금", "연금"],
    "건강보험": ["건강보험", "건 강", "건보"],
    "장기요양": ["장기요양", "요양보험", "장기 요양"],
    "고용보험": ["고용보험", "고용 보험"],
    "공제총액": ["공제총액", "공제계", "공제합계", "공제 계", "공제금액"],
    "실수령":   ["실수령", "실지급", "차인지급", "차감지급", "실지급액", "차인지급액", "실수령액"],
}
KEY_ORDER = list(FIELDS.keys())
MONTH_RE = re.compile(r'(\d{1,2})\s*월|(\d{2})[.\-]?(\d{1,2})|\b(1[0-2]|0?[1-9])월?\b')


def find_header(ws, max_scan=18):
    """헤더행 탐색: 핵심 필드가 가장 많이 매칭되는 행. (행번호, 매칭필드set) 반환."""
    best_row, best_fields = None, set()
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan, max_col=50, values_only=True))
    for i, row in enumerate(rows):
        cells = [str(c) for c in row if c is not None]
        joined = " ".join(cells)
        found = set()
        for canon, syns in FIELDS.items():
            for s in syns:
                if s in joined:
                    found.add(canon)
                    break
        if len(found) > len(best_fields):
            best_fields, best_row = found, i + 1
    return best_row, best_fields


def is_month_sheet(name):
    return bool(re.search(r'\d{1,2}\s*월|^\d{2}[.\-]?\d{1,2}|월$', name.strip()))


def main():
    cls_path = os.path.join(OUT_DIR, "file_classification.json")
    with open(cls_path, encoding="utf-8") as f:
        records = json.load(f)
    targets = [r for r in records
               if r["label"] == "급여대장" and r["ext"] in (".xlsx", ".xlsm")]

    import openpyxl
    results = []
    fail = 0
    for r in targets:
        path = os.path.join(DATA_ROOT, r["path"])
        info = {"path": r["path"], "handler": r["handler"], "sheets": 0,
                "month_tabs": 0, "ledger_sheets": 0, "fields": [], "err": None}
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            info["sheets"] = len(wb.sheetnames)
            info["month_tabs"] = sum(1 for s in wb.sheetnames if is_month_sheet(s))
            union_fields = set()
            ledger_sheets = 0
            for ws in wb.worksheets[:20]:
                hr, ff = find_header(ws)
                if len(ff) >= 4:           # 급여대장 시트로 인정할 최소 필드 수
                    ledger_sheets += 1
                    union_fields |= ff
            info["ledger_sheets"] = ledger_sheets
            info["fields"] = sorted(union_fields, key=lambda x: KEY_ORDER.index(x))
            wb.close()
        except Exception as e:
            info["err"] = f"{type(e).__name__}"
            fail += 1
        results.append(info)

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "profile.json"), "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=1)

    # ── 군집: 핵심 필드 존재 패턴(11비트) ──
    ok = [r for r in results if not r["err"] and r["ledger_sheets"] > 0]
    noledger = [r for r in results if not r["err"] and r["ledger_sheets"] == 0]
    errs = [r for r in results if r["err"]]

    def sig(r):
        return "".join("1" if k in r["fields"] else "0" for k in KEY_ORDER)

    by_sig = defaultdict(list)
    for r in ok:
        by_sig[sig(r)].append(r)

    # 필드 개수별 분포
    fieldcount = Counter(len(r["fields"]) for r in ok)
    # 월탭 분포
    multi_month = [r for r in ok if r["month_tabs"] >= 2]
    single = [r for r in ok if r["month_tabs"] < 2]
    # 필드별 등장률
    field_freq = Counter()
    for r in ok:
        for fld in r["fields"]:
            field_freq[fld] += 1

    lines = []
    lines.append(f"대상 급여대장 엑셀: {len(targets)}")
    lines.append(f"  헤더 인식 성공(급여대장 시트 1개+): {len(ok)}")
    lines.append(f"  엑셀은 열렸으나 급여대장 헤더 못찾음: {len(noledger)}")
    lines.append(f"  열기 실패(손상 등): {len(errs)}")
    lines.append("")
    lines.append("[핵심필드 개수별 분포] (11개 만점)")
    for cnt, n in sorted(fieldcount.items(), reverse=True):
        lines.append(f"  {cnt}개 필드: {n}건")
    lines.append("")
    lines.append(f"[월탭 구조] 단일/소수월 {len(single)} / 연간누적형(2탭+) {len(multi_month)}")
    lines.append("")
    lines.append("[필드별 등장률]")
    for fld in KEY_ORDER:
        n = field_freq[fld]
        pct = 100 * n // max(len(ok), 1)
        lines.append(f"  {fld}: {n} ({pct}%)")
    lines.append("")
    lines.append(f"[구조 시그니처 군집] (상위 12개 / 총 {len(by_sig)}종)")
    lines.append(f"  (비트순서: {' '.join(KEY_ORDER)})")
    top = sorted(by_sig.items(), key=lambda kv: len(kv[1]), reverse=True)[:12]
    for s, group in top:
        present = [KEY_ORDER[i] for i, b in enumerate(s) if b == "1"]
        avg_month = sum(g["month_tabs"] for g in group) / len(group)
        example = group[0]["path"].split("\\")[-1] if "\\" in group[0]["path"] else group[0]["path"].split("/")[-1]
        lines.append(f"  [{len(group)}건] {s}  필드={present}")
        lines.append(f"          예: {example}")

    summary = "\n".join(lines)
    with open(os.path.join(OUT_DIR, "profile_summary.txt"), "w", encoding="utf-8") as f:
        f.write(summary)
    print(summary)


if __name__ == "__main__":
    main()
