# -*- coding: utf-8 -*-
"""
하네스 1차 - 목표 1: 골든 파서 v1 (가로 표형 급여대장 → 직원별 dict)
- 급여대장 시트 자동 선택 + 2~3줄 병합 헤더 평탄화 + 컬럼 자동 인식.
- 주민번호는 '데이터 행 판별'에만 사용하고 절대 저장하지 않음(성명만).
- v1 범위: 가로 표형(한 줄 = 직원). 명세서(세로)형·일용 블록형은 다음 확장.
- 테스트 모드: `python harness/parser_v1.py --test` → 샘플 몇 개를 뽑아 결과 출력.
"""
import os, json, re, sys, warnings
warnings.filterwarnings("ignore")

DATA_ROOT = os.environ.get(
    "PAYROLL_DATA_ROOT",
    r"C:\Users\fair0\OneDrive\바탕 화면\급여아웃소싱 서류들",
)
OUT_DIR = os.path.join(DATA_ROOT, "_harness_out")

FIELDS = {
    "성명":     ["성명", "이름", "성 명", "직원명"],
    "기본급":   ["기본급", "기본 급"],
    "과세총액": ["과세총액", "과세대상액", "과세대상", "과세소득금액", "과세소득", "과세금액",
                "과세계", "과세합계", "총과세", "보수총액", "보수월액", "과세보수", "과세임금",
                "임금총액", "노무비총액", "노임총액"],  # 일용 대장: 총액=과세 기반(고용보험=총액×요율 실측)
    # 주의: 지방세를 소득세보다 먼저 — '지방소득세' 헤더가 소득세로 선점되는 것 방지
    "지방세":   ["지방소득세", "지방세", "주민세"],
    "소득세":   ["소득세", "갑근세"],
    "국민연금": ["국민연금"],
    "건강보험": ["건강보험"],
    "장기요양": ["장기요양", "요양보험"],
    "고용보험": ["고용보험"],
    # 연말/중도정산 정산액(부호 그대로: 음수=환급, 양수=추가징수). 공제란에 표시.
    "연말정산": ["연말정산", "중도정산", "연말정산세액", "연말정산정산액"],
    "공제총액": ["공제총액", "공제계", "공제합계", "공제금액", "공제 계"],
    "실수령":   ["차인지급액", "차인지급", "실지급액", "실수령액", "실수령", "실지급", "차감지급", "실지급총액"],
    "지급총액": ["지급총액", "지급합계", "지급계", "급여계", "총지급액", "지급액계",
                "지급 합계", "지급액 계", "지급액", "총지급"],
    # 일용 대장 전용(명세서·신고 산출용) — 새 필드
    "일당":     ["일당", "노무비단가", "노임단가"],
    "근무일수": ["출력일수", "총근무일수", "출역일수", "근무일수"],
    # 시급제 일용(제이앤드씨류): 시급×총근무시간, 평균시간=신고서의 '일평균 근로시간'
    "시급":     ["시급"],
    "근무시간": ["총근무시간", "총근로시간"],
    "평균시간": ["평균시간", "일평균근로시간"],
}

# 짧아서 다른 항목(기타공제액·연차일수 등)과 오인되기 쉬운 라벨:
# 공백 제거한 헤더가 '정확히' 이 값일 때만 매핑.
FIELDS_EXACT = {"공제액": "공제총액", "공수": "근무일수", "일수": "근무일수", "단가": "일당"}

# 부정어 가드: 헤더에 이 단어가 있으면 해당 필드로 매핑 금지
# (기지급액→지급총액, 시트제목 '급여 계산'→급여계 오인 방지)
# 주의: '비과세'는 여기 넣지 말 것 — 삼산회관류는 '비과세계' 라벨에 실제 과세총액이
# 들어있음(실측: 기본급+수당=해당열, 엔진 고용보험 검증 96~100%). 대신 아래
# 2단계 바인딩으로 '비과세 아닌 과세계 열'이 있으면 그쪽을 우선.
FIELDS_NEG = {
    "지급총액": ("기지급", "미지급", "선지급", "가지급", "계산"),
    "공제총액": ("지각공제", "결근공제"),
}


def _nows(s):
    """공백·줄바꿈 전부 제거(일용 대장 '노무비\\n총  액' 류 헤더 대응)."""
    return re.sub(r"\s+", "", str(s)) if s is not None else ""

# 기타공제 — 명확한 항목명만(과다합산 방지). '공제'·'차감' 같은 광범위 키워드 제외.
EXTRA_DED_PAT = ["상조", "조합비", "사우회", "경조", "학자금", "가불", "기숙사비", "친목"]
JUMIN_RE = re.compile(r'\d{6}\s*[-]?\s*\d{6,7}')
NAME_RE = re.compile(r'^[가-힣]{2,4}$')
NUM_RE = re.compile(r'^-?[\d,]+(\.\d+)?$')

# 성명 칸으로 오인되기 쉬운 직급·직위·부서·기타 단어(이름으로 채택 금지)
NAME_BLOCK = {
    "사원", "조리원", "실장", "주임", "점장", "영양사", "팀장", "시급직", "주방", "홀",
    "매니저", "원장", "간호사", "간호조무사", "반장", "공장장", "대표", "이사", "부장",
    "과장", "차장", "대리", "전무", "상무", "사장", "회장", "직원", "근로자", "성명",
    "이름", "합계", "소계", "총계", "기숙사비", "베이커리", "파트", "알바", "정규직",
    "계약직", "일용직", "월급직", "관리자", "사무", "생산", "배송", "판매", "본사", "지점",
    "임원", "전장", "부장님", "실장님", "점주", "사원명", "직급", "직위", "부서", "성함",
}


def is_name(s):
    """한글 2~4자이고 직급/비이름 목록에 없으면 성명으로 인정."""
    s = (s or "").strip()
    return bool(NAME_RE.match(s)) and s not in NAME_BLOCK

# 급여대장 시트로 볼 이름 힌트 / 제외할 시트
SHEET_GOOD = ["급여대장", "급상여", "급여", "임금대장", "명세"]
SHEET_BAD = ["세율", "세액표", "base", "data", "설명", "근로기준법", "사업자", "연차", "안내"]


def flatten_cols(rows, col_count):
    """여러 헤더행을 컬럼별로 세로 결합."""
    out = []
    for c in range(col_count):
        parts = []
        for row in rows:
            v = row[c] if c < len(row) else None
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        out.append(" ".join(parts))
    return out


def match_field(header_text):
    """공백·줄바꿈 무시 매칭(사전 순서 = 우선순위) + 부정어 가드 + 정확일치 라벨."""
    ht = _nows(header_text)
    if not ht:
        return None
    for canon, syns in FIELDS.items():
        if any(p in ht for p in FIELDS_NEG.get(canon, ())):
            continue
        for s in syns:
            ns = _nows(s)
            if ns and ns in ht:
                return canon
    return FIELDS_EXACT.get(ht)


def parse_num(v):
    if v is None:
        return None
    s = str(v).replace(",", "").replace(" ", "").strip()
    if s in ("", "-", "—"):
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return None


def score_sheet(ws):
    """시트가 가로 표형 급여대장인지 점수화. (헤더행, 필드매핑, 데이터시작행) or None.
    주민번호가 있으면 그 행을, 없으면 헤더(성명+급여항목) 다음 행을 데이터 시작으로 본다."""
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), max_col=min(ws.max_column, 60), values_only=True))
    if not rows:
        return None

    def field_hits(row):
        joined = " ".join(str(c) for c in row if c is not None)
        f = set()
        for canon, syns in FIELDS.items():
            # 보조필드는 점수에서 제외 — 상용/일용 표가 쌓인 시트에서
            # 일용 헤더가 상용 헤더를 이겨 상용 직원이 잘림(헤더 선택은 핵심필드만)
            if canon in ("일당", "근무일수", "시급", "근무시간", "평균시간"):
                continue
            if any(s in joined for s in syns):
                f.add(canon)
        return f

    # 1) 후보 헤더행 = 핵심필드가 가장 많은 행
    #    (헤더가 2~3줄로 쪼개진 경우가 많아, 단일행 문턱은 낮게 두고
    #     최종 판정은 아래 '합쳐진 헤더(colmap)' 기준으로 한다)
    best_i, best_f = None, set()
    for i, row in enumerate(rows):
        f = field_hits(row)
        if len(f) > len(best_f):
            best_f, best_i = f, i
    if best_i is None or len(best_f) < 2:
        return None

    # 2) 데이터 시작행: 헤더 이후 첫 주민번호행, 없으면 헤더+1
    data_start = None
    for i in range(best_i, len(rows)):
        joined = " ".join(str(c) for c in rows[i] if c is not None)
        if JUMIN_RE.search(joined):
            data_start = i
            break
    if data_start is None or data_start <= best_i:
        data_start = best_i + 1

    # 3) 헤더 결합(2~3줄 병합 대응) → 컬럼→필드 매핑
    hstart = max(0, min(best_i, data_start - 1) - 2)
    header_rows = rows[hstart:data_start]
    col_count = max(len(r) for r in rows)
    flat = flatten_cols(header_rows, col_count)
    colmap = {}
    prov = {}  # 께름칙한 매칭(예: '비과세계'→과세총액) — 더 나은 열 없을 때만 채택
    for ci, htext in enumerate(flat):
        f = match_field(htext)
        if not f or f in colmap.values():
            continue
        if f == "과세총액" and "비과세" in _nows(htext):
            prov.setdefault(f, ci)
            continue
        colmap[ci] = f
    for f, ci in prov.items():
        if f not in colmap.values():
            colmap[ci] = f
    fields = set(colmap.values())
    if len(fields) < 4:
        return None
    # 기타공제 후보 컬럼(핵심필드 미매핑 + 공제성 키워드)
    extra_ded = [ci for ci, h in enumerate(flat)
                 if ci not in colmap and h and any(p in h for p in EXTRA_DED_PAT)]
    # 일용 대장 여부: 총액=과세=지급 동일 구조(임금총액·노무비·출역 헤더)
    daily = any(k in _nows(h) for h in flat if h
                for k in ("임금총액", "노무비", "출역"))
    return {"data_start": data_start, "header_rows": (hstart, data_start),
            "colmap": colmap, "fields": fields, "col_count": col_count,
            "extra_ded": extra_ded, "daily": daily}


def pick_and_parse(wb):
    """워크북에서 급여대장 시트들을 골라 파싱. 시트별 결과 리스트."""
    out = []
    for ws in wb.worksheets:
        nm = ws.title.lower()
        if any(b in nm for b in SHEET_BAD):
            continue
        sc = score_sheet(ws)
        if not sc:
            continue
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=sc["col_count"], values_only=True))
        colmap = sc["colmap"]
        daily = sc.get("daily") or any(k in ws.title for k in ("일용", "노임", "노무"))
        # 성명 컬럼: 헤더매핑 우선, 없으면 데이터에서 한글이름 최빈 컬럼 추정
        name_col = next((c for c, f in colmap.items() if f == "성명"), None)
        employees = []
        for row in rows[sc["header_rows"][1]:]:
            joined = " ".join(str(c) for c in row if c is not None)
            # 합계행 등 제외
            if any(k in joined for k in ["합계", "총계", "소계", "합 계"]):
                continue
            has_jumin = bool(JUMIN_RE.search(joined))
            # 성명 후보 (직급·비이름 단어는 건너뜀)
            nm = None
            if name_col is not None and name_col < len(row) and row[name_col]:
                cand = str(row[name_col]).strip()
                if is_name(cand):
                    nm = cand
            if nm is None:
                # 직급 등 제외하고 행에서 진짜 이름 탐색(주민번호 있을 때 우선)
                for c in row:
                    if c and is_name(str(c).strip()):
                        nm = str(c).strip(); break
            if not nm:
                continue
            emp = {"성명": nm}
            for c, f in colmap.items():
                if f == "성명":
                    continue
                if c < len(row):
                    val = parse_num(row[c])
                    if val is not None:
                        emp[f] = val
            # 기타공제 합산(상조·가불·정산 등 — 공제총액 정합성용)
            extra = 0
            for c in sc.get("extra_ded", []):
                if c < len(row):
                    v = parse_num(row[c])
                    if v is not None:
                        extra += v
            if extra:
                emp["기타공제"] = extra
            # 일용 대장: 총액 칸 하나(임금총액·노무비총액)가 과세총액으로 잡힘
            # → 지급총액 칸이 따로 없으므로 동일값 채움(일용은 총액=과세=지급)
            if daily and emp.get("과세총액") is not None:
                emp.setdefault("지급총액", emp["과세총액"])
            if len(emp) >= 3:  # 성명 + 숫자필드 2개+
                employees.append(emp)
        if employees:
            out.append({"sheet": ws.title, "fields": sorted(sc["fields"]),
                        "n_emp": len(employees), "employees": employees})
    return out


def extract(path):
    import openpyxl
    res = {"path": os.path.relpath(path, DATA_ROOT), "ok": False, "sheets": [], "err": None}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = pick_and_parse(wb)
        wb.close()
        res["sheets"] = sheets
        res["ok"] = len(sheets) > 0
        res["n_emp_total"] = sum(s["n_emp"] for s in sheets)
    except Exception as e:
        res["err"] = f"{type(e).__name__}: {e}"
    return res


def test_mode():
    prof = json.load(open(os.path.join(OUT_DIR, "profile.json"), encoding="utf-8"))
    keys = ["가람떡집", "일등한우(주)", "니쿠미야", "선우기술", "파보네", "제스트"]
    picked = []
    for k in keys:
        for r in prof:
            if k in r["path"] and r["path"].endswith(".xlsx") and r not in picked:
                picked.append(r); break
    lines = []
    for r in picked:
        res = extract(os.path.join(DATA_ROOT, r["path"]))
        base = r["path"].split("\\")[-1] if "\\" in r["path"] else r["path"].split("/")[-1]
        lines.append(f"\n#### {base}")
        if res["err"]:
            lines.append("   ERR: " + res["err"]); continue
        if not res["ok"]:
            lines.append("   [실패] 급여대장 표 인식 실패"); continue
        for s in res["sheets"][:2]:
            lines.append(f"   시트'{s['sheet']}' 필드{s['fields']} 직원{s['n_emp']}명")
            for e in s["employees"][:3]:
                lines.append("     - " + str({k: v for k, v in e.items()}))
    out = "\n".join(lines)
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "parser_test.txt"), "w", encoding="utf-8") as f:
        f.write(out)
    try:
        print(out)
    except UnicodeEncodeError:
        print("(콘솔 인코딩 문제로 생략 — parser_test.txt 참조)")


# ══════════════════════════════════════════════════════════════
#  세로 명세서형 파서 (에스에스테크·화인마취류)
#  지급항목|금액 / 공제항목|금액 세로배치. 주민번호 없이 라벨로 인식.
# ══════════════════════════════════════════════════════════════
PAY_LABEL = ["지급항목", "지급내역", "지 급 항 목"]
DED_LABEL = ["공제항목", "공제내역", "공 제 항 목"]
NAME_LABEL = ["성명", "성 명", "성  명", "성    명", "직원명", "근로자명"]
FIELDS_VERT = dict(FIELDS)
FIELDS_VERT["지급총액"] = ["지급총액", "지급 총액", "지급계", "지급합계", "지 급 총 액"]


def _cellstr(v):
    return "" if v is None else str(v).strip()


def _norm(s):
    """공백(자간 벌리기 포함) 전부 제거해 라벨 매칭."""
    return re.sub(r"\s+", "", str(s)) if s is not None else ""

NAME_KEYS = {"성명", "사원명", "직원명", "근로자명", "성함", "이름", "사원성명", "근로자성명"}
# 공백 제거한 필드 동의어(성명 제외)
FIELDS_VN = {k: [_norm(s) for s in v] for k, v in FIELDS_VERT.items() if k != "성명"}


DED_FIELDS = ["소득세", "지방세", "국민연금", "건강보험", "장기요양", "고용보험", "공제총액"]


def _value_right(ws, r, c, cmax):
    """라벨(r,c) 오른쪽으로 스캔 → 첫 숫자 반환. 다른 라벨(글자) 만나면 멈춤(None).
    병합 셀로 값이 여러 칸 떨어진 경우 대응."""
    for cc in range(c + 1, cmax + 1):
        v = ws.cell(row=r, column=cc).value
        if v is None:
            continue
        n = parse_num(v)
        if n is not None:
            return n
        return None  # 숫자 아닌 텍스트 = 다음 라벨 → 멈춤
    return None


def _validate_emp(emp):
    """상식 검증: 값이 말이 되는지. 틀리면 False(→ 사람 확인 대상)."""
    base = emp.get("지급총액") or emp.get("기본급") or emp.get("과세총액")
    if base is None or base <= 0:
        return False  # 기준 급여 없으면 신뢰 불가
    for d in DED_FIELDS:
        if d in emp and emp[d] > base:   # 공제가 급여보다 큼 = 오독
            return False
    if "실수령" in emp:
        if emp["실수령"] <= 0 or emp["실수령"] > base * 1.5:
            return False
    return True


def extract_vertical(path):
    """세로 명세서형: 공백무시 + '사원명' 앵커 + 병합셀 값읽기 + 상식검증.
    검증 통과분만 employees에 담고, 실패분 수는 n_flagged로 보고(틀린 숫자 유입 차단)."""
    import openpyxl
    res = {"path": os.path.relpath(path, DATA_ROOT), "ok": False, "sheets": [],
           "err": None, "mode": "vertical", "n_flagged": 0}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)  # 병합정보 위해 non-read_only
    except Exception as e:
        res["err"] = f"{type(e).__name__}: {e}"
        return res
    flagged = 0
    for ws in wb.worksheets:
        low = ws.title.lower()
        if any(b in low for b in SHEET_BAD):
            continue
        maxr, maxc = min(ws.max_row, 250), min(ws.max_column, 40)
        # 1) 사원명 앵커
        anchors = []
        for r in range(1, maxr + 1):
            for c in range(1, maxc + 1):
                if _norm(ws.cell(row=r, column=c).value) in NAME_KEYS:
                    nm = None
                    for k in range(c + 1, min(c + 4, maxc + 1)):
                        s = ws.cell(row=r, column=k).value
                        if s and str(s).strip() and _norm(s) not in NAME_KEYS:
                            cand = str(s).strip()
                            if (re.match(r"^[가-힣A-Za-z]", cand) and len(cand) <= 12
                                    and cand not in NAME_BLOCK):
                                nm = cand
                            break
                    if nm:
                        anchors.append((r, c, nm))
        if not anchors:
            continue
        acols = sorted(set(a[1] for a in anchors))
        emps, seen = [], set()
        for (ar, ac, nm) in anchors:
            nexts = [x for x in acols if x > ac + 2]
            cmax = (min(nexts) - 1) if nexts else maxc
            below = [a[0] for a in anchors if a[0] > ar and abs(a[1] - ac) <= 2]
            rmax = min(below) if below else min(ar + 22, maxr)
            emp = {"성명": nm}
            for canon, syns in FIELDS_VN.items():
                got = None
                for r in range(ar, rmax + 1):
                    for c in range(ac, cmax + 1):
                        cell = _norm(ws.cell(row=r, column=c).value)
                        if cell and any(s and s in cell for s in syns):
                            got = _value_right(ws, r, c, cmax)
                            if got is not None:
                                break
                    if got is not None:
                        break
                if got is not None:
                    emp[canon] = got
            if len(emp) < 3:
                continue
            if not _validate_emp(emp):
                flagged += 1
                continue
            key = (nm, tuple(sorted((k, v) for k, v in emp.items() if k != "성명")))
            if key in seen:          # 완전 동일 레코드 중복 제거
                continue
            seen.add(key)
            emps.append(emp)
        if emps:
            fields = sorted({k for e in emps for k in e})
            res["sheets"].append({"sheet": ws.title, "fields": fields, "n_emp": len(emps), "employees": emps})
    wb.close()
    res["n_flagged"] = flagged
    res["ok"] = len(res["sheets"]) > 0
    res["n_emp_total"] = sum(s["n_emp"] for s in res["sheets"])
    return res


def extract_any(path):
    """가로형(검증됨) 우선, 실패 시 세로 명세서형(상식검증 통과분만 채택).
    세로형은 값이 의심스러우면 채택하지 않고 n_flagged로만 집계 → 틀린 숫자 유입 차단."""
    r = extract(path)
    if r["ok"]:
        r["mode"] = "horizontal"
        return r
    v = extract_vertical(path)
    if v["ok"]:
        return v
    # 세로형도 실패(플래그만 있음)면 그 정보 보존
    r["mode"] = "none"
    r["n_flagged"] = v.get("n_flagged", 0)
    return r


def run_all():
    """전체 급여대장 엑셀에 파서 적용 → 커버리지 산출."""
    from collections import Counter
    cls = json.load(open(os.path.join(OUT_DIR, "file_classification.json"), encoding="utf-8"))
    targets = [r for r in cls if r["label"] == "급여대장" and r["ext"] in (".xlsx", ".xlsm")]

    # 모수 종류 분류: 비급여(오분류) / 일용(확장대상) / 상용(가로표 기대)
    NONLEDGER = ["연차", "명부", "근로자정보", "직원정보", "직원 정보", "정보(", "양식",
                 "근무표", "피부양", "취득", "상실", "주소록", "수습", "명단"]
    def kind(rel):
        b = rel.split("\\")[-1] if "\\" in rel else rel.split("/")[-1]
        if any(k in b for k in NONLEDGER):
            return "비급여(오분류)"
        if "일용" in rel:
            return "일용직"
        return "상용"
    for r in targets:
        r["_kind"] = kind(r["path"])

    results = []
    for r in targets:
        rr = extract_any(os.path.join(DATA_ROOT, r["path"]))
        rr["_kind"] = r["_kind"]
        results.append(rr)

    ok = [r for r in results if r["ok"]]
    err = [r for r in results if not r["ok"] and r["err"]]
    norec = [r for r in results if not r["ok"] and not r["err"]]
    total_emp = sum(r.get("n_emp_total", 0) for r in ok)
    total_sheets = sum(len(r["sheets"]) for r in ok)
    mode_cnt = Counter(r.get("mode", "?") for r in ok)

    # 필드 채움률(성공 파일 기준, 시트 union)
    field_files = Counter()
    for r in ok:
        fs = set()
        for s in r["sheets"]:
            fs |= set(s["fields"])
        for f in fs:
            field_files[f] += 1

    # 결과 저장(개인정보 포함 → 자료 쪽)
    with open(os.path.join(OUT_DIR, "parser_output.json"), "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False)

    lines = []
    lines.append(f"대상 급여대장 엑셀(.xlsx/.xlsm): {len(targets)}")
    lines.append(f"  파싱 성공(직원 1명+ 추출): {len(ok)}  ({100*len(ok)//max(len(targets),1)}%)")
    lines.append(f"    - 가로 표형: {mode_cnt.get('horizontal',0)}  / 세로 명세서형(검증통과): {mode_cnt.get('vertical',0)}")
    lines.append(f"  세로형 상식검증 탈락(사람 확인 대상) 레코드: {sum(r.get('n_flagged',0) for r in results)}")
    lines.append(f"  표 인식 실패(열림·표없음): {len(norec)}")
    lines.append(f"  열기 실패(손상 등): {len(err)}")
    lines.append(f"  추출 직원 레코드 총: {total_emp}")
    lines.append(f"  추출 급여대장 시트(월탭 분해 포함) 총: {total_sheets}")
    lines.append("")
    lines.append("[모수 종류별 성공률] — 진짜 성공률은 '상용' 기준")
    for kd in ["상용", "일용직", "비급여(오분류)"]:
        grp = [r for r in results if r.get("_kind") == kd]
        gok = [r for r in grp if r["ok"]]
        lines.append(f"  {kd}: {len(gok)}/{len(grp)} 성공 ({100*len(gok)//max(len(grp),1)}%)")
    lines.append("")
    lines.append("[성공 파일의 필드 포함률]")
    for f in FIELDS:
        n = field_files[f]
        lines.append(f"  {f}: {n} ({100*n//max(len(ok),1)}%)")
    lines.append("")
    lines.append("[실패 파일 예시 20]")
    for r in (norec + err)[:20]:
        base = r["path"].split("\\")[-1] if "\\" in r["path"] else r["path"].split("/")[-1]
        lines.append(f"  {'ERR:'+r['err'][:30] if r['err'] else '표없음'} | {base}")

    out = "\n".join(lines)
    with open(os.path.join(OUT_DIR, "parser_coverage.txt"), "w", encoding="utf-8") as f:
        f.write(out)
    # 콘솔이 cp949면 특수문자(—·⚠ 등)에서 죽으므로 인코딩 불가 문자는 버리고 출력(파일은 정상).
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    try:
        print(out)
    except UnicodeEncodeError:
        print("(coverage printed to parser_coverage.txt)")


if __name__ == "__main__":
    if "--test" in sys.argv:
        test_mode()
    elif "--all" in sys.argv:
        run_all()
    else:
        print("사용: --test (샘플 검증) | --all (전체 커버리지)")
