# -*- coding: utf-8 -*-
"""
하네스 1차 - 목표 5: 근태↔대장 매칭 키 실태 조사
- 근태 파일(분류 결과)에서 ① 성명 목록 ② 보조키 후보(사번·부서·직급·입사일·연락처·주민번호 존재여부)를 읽고
- 같은 사업장 급여대장 성명과 대조 → 사업장별 3분류:
  · 이름 단독 매칭 가능(동명이인 없음 + 근태·대장 이름 겹침 양호)
  · 보조키 필요(동명이인 존재 or 이름 표기 차이 큼)
  · 매칭 불가/자료부족(근태 이름 추출 실패 등)
- 주민번호는 '존재 여부'만 기록, 값은 절대 추출·저장하지 않음.
- 결과: _harness_out/matching_summary.txt
"""
import os, json, re, warnings
from collections import Counter, defaultdict
warnings.filterwarnings("ignore")

DATA_ROOT = os.environ.get(
    "PAYROLL_DATA_ROOT",
    r"C:\Users\fair0\OneDrive\바탕 화면\급여아웃소싱 서류들",
)
OUT_DIR = os.path.join(DATA_ROOT, "_harness_out")
WRAP = {"급여", "1. 급여", "급여자료", "기타", "급여관리"}

NAME_RE = re.compile(r'^[가-힣]{2,4}$')
JUMIN_RE = re.compile(r'\d{6}\s*[-]?\s*\d{6,7}')
NAME_BLOCK = {
    "사원", "조리원", "실장", "주임", "점장", "영양사", "팀장", "시급직", "주방", "홀",
    "매니저", "원장", "간호사", "간호조무사", "반장", "공장장", "대표", "이사", "부장",
    "과장", "차장", "대리", "전무", "상무", "사장", "회장", "직원", "근로자", "성명",
    "이름", "합계", "소계", "총계", "휴무", "결근", "지각", "조퇴", "연장", "야간",
    "주간", "출근", "퇴근", "휴가", "월차", "연차", "반차", "특근", "일용", "상용",
}
AUX_KEYS = {
    "사번": ["사번", "사원번호", "사원 번호", "직원번호"],
    "부서": ["부서", "소속", "부 서"],
    "직급": ["직급", "직위", "직 급"],
    "입사일": ["입사일", "입사 일", "입사년월", "입사"],
    "연락처": ["연락처", "전화", "휴대", "핸드폰", "H.P", "HP"],
}


def raw_site(rel):
    parts = rel.replace("/", "\\").split("\\")
    rest = parts[1:]
    i = 0
    while i < len(rest) and rest[i] in WRAP:
        i += 1
    if i >= len(rest):
        return None
    s = re.sub(r'^\d+\s*[.\-]\s*', '', rest[i]).strip()
    return s or None


def site_base(s):
    if not s:
        return s
    return re.sub(r'\s*[\(\[].*?[\)\]]\s*$', '', s).strip() or s


def is_name(s):
    s = (s or "").strip()
    return bool(NAME_RE.match(s)) and s not in NAME_BLOCK


def scan_attendance(path):
    """근태 파일에서 (이름집합, 보조키존재셋, 주민번호존재). 이름은 성명 헤더 컬럼 우선, 없으면 휴리스틱."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, set(), False
    names, aux, jumin = set(), set(), False
    try:
        for ws in wb.worksheets[:8]:
            rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200),
                                     max_col=min(ws.max_column, 40), values_only=True))
            if not rows:
                continue
            # 보조키·주민 존재 검사(상단 12행)
            for row in rows[:12]:
                for c in row:
                    if c is None:
                        continue
                    s = str(c)
                    if JUMIN_RE.search(s):
                        jumin = True
                    ns = re.sub(r'\s+', '', s)
                    for k, syns in AUX_KEYS.items():
                        if any(re.sub(r'\s+', '', x) in ns for x in syns):
                            aux.add(k)
            # 성명 헤더 컬럼 찾기
            name_cols = []
            header_row_i = None
            for i, row in enumerate(rows[:12]):
                for ci, c in enumerate(row):
                    ns = re.sub(r'\s+', '', str(c)) if c is not None else ""
                    if ns in ("성명", "이름", "사원명", "근로자명", "성함"):
                        name_cols.append(ci)
                        header_row_i = i
            if name_cols and header_row_i is not None:
                for row in rows[header_row_i + 1:]:
                    for ci in name_cols:
                        if ci < len(row) and row[ci] and is_name(str(row[ci]).strip()):
                            names.add(str(row[ci]).strip())
            else:
                # 휴리스틱: 주민번호와 같은 행의 이름 셀만(오탐 억제)
                for row in rows:
                    joined = " ".join(str(c) for c in row if c is not None)
                    if JUMIN_RE.search(joined):
                        jumin = True
                        for c in row:
                            if c and is_name(str(c).strip()):
                                names.add(str(c).strip())
                                break
    finally:
        wb.close()
    return names, aux, jumin


def main():
    cls = json.load(open(os.path.join(OUT_DIR, "file_classification.json"), encoding="utf-8"))
    res = json.load(open(os.path.join(OUT_DIR, "parser_output.json"), encoding="utf-8"))

    # 대장 이름/동명이인(시트 내 중복) per site
    ledger_names = defaultdict(set)
    ledger_dupe = defaultdict(set)
    for r in res:
        if not r["ok"]:
            continue
        sb = site_base(raw_site(r["path"]))
        for s in r["sheets"]:
            ns = [e["성명"] for e in s["employees"] if e.get("성명")]
            ledger_names[sb].update(ns)
            cnt = Counter(ns)
            rep = sum(1 for c in cnt.values() if c >= 2)
            if len(cnt) and rep / len(cnt) <= 0.3:
                for nm, c in cnt.items():
                    if c >= 2:
                        ledger_dupe[sb].add(nm)

    # 근태 파일 스캔 per site
    att = [r for r in cls if r["label"] == "근태" and r["ext"] in (".xlsx", ".xlsm")]
    att_names = defaultdict(set)
    att_aux = defaultdict(set)
    att_jumin = defaultdict(bool)
    att_cnt = Counter()
    for r in att:
        sb = site_base(raw_site(r["path"]))
        if not sb:
            continue
        att_cnt[sb] += 1
        names, aux, jm = scan_attendance(os.path.join(DATA_ROOT, r["path"]))
        if names:
            att_names[sb].update(names)
        att_aux[sb].update(aux)
        att_jumin[sb] = att_jumin[sb] or jm

    # 사업장별 판정
    lines = []
    lines.append("=" * 50)
    lines.append("목표5. 근태-대장 매칭 키 실태 조사")
    lines.append("=" * 50)
    lines.append(f"근태 엑셀 파일: {len(att)}개 / 근태 보유 사업장: {len(att_cnt)}곳")
    lines.append("")
    verdict = Counter()
    both = sorted(set(att_names) & set(ledger_names), key=lambda s: -len(att_names[s]))
    for sb in both:
        an, ln = att_names[sb], ledger_names[sb]
        inter = an & ln
        only_att = sorted(an - ln)
        overlap = 100 * len(inter) // max(len(an), 1)
        dupes = ledger_dupe.get(sb, set())
        aux = sorted(att_aux.get(sb, set()))
        if dupes:
            v = "보조키 필요(동명이인)"
        elif overlap >= 70:
            v = "이름 단독 매칭 가능"
        elif overlap >= 30:
            v = "보조키 필요(표기차이)"
        else:
            v = "매칭 불가/자료부족"
        verdict[v] += 1
        lines.append(f"[{sb}] {v}")
        lines.append(f"   근태이름 {len(an)}명, 대장이름 {len(ln)}명, 겹침 {len(inter)}명({overlap}%)")
        lines.append(f"   근태 보조키: {aux if aux else '없음'} / 주민번호 존재: {'예' if att_jumin[sb] else '아니오'}")
        if dupes:
            lines.append(f"   대장 동명이인: {sorted(dupes)}")
        if only_att[:8]:
            lines.append(f"   근태에만 있는 이름(unmatched) {len(only_att)}명 예: {only_att[:8]}")
    # 근태만 있고 대장 파싱 없는 곳 / 대장만 있는 곳
    att_only = sorted(set(att_cnt) - set(ledger_names))
    lines.append("")
    lines.append("[판정 분포]")
    for k, v in verdict.most_common():
        lines.append(f"  {k}: {v}곳")
    if att_only:
        lines.append(f"\n근태는 있으나 대장 파싱결과 없는 사업장 {len(att_only)}곳: " +
                     ", ".join(att_only[:12]) + (" ..." if len(att_only) > 12 else ""))
    no_names = sorted(s for s in att_cnt if s in ledger_names and not att_names.get(s))
    if no_names:
        lines.append(f"근태파일에서 이름 추출 실패 {len(no_names)}곳: " + ", ".join(no_names[:12]))
    lines.append("\n※ 주민번호는 존재 여부만 기록(값 미추출). 동명이인 사업장은 이름만으로 매칭 금지(설계 D2).")

    out = "\n".join(lines)
    with open(os.path.join(OUT_DIR, "matching_summary.txt"), "w", encoding="utf-8") as f:
        f.write(out)
    try:
        print(out[:1200])
    except UnicodeEncodeError:
        print("(matching_summary.txt 참조)")


if __name__ == "__main__":
    main()
